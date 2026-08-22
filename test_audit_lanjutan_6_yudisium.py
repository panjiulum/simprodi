# -*- coding: utf-8 -*-
"""
test_audit_lanjutan_6_yudisium.py — Uji Pengembangan Lanjutan 6 (temuan
tambahan, laporan pengguna): dampak bug nilai_angka_ke_huruf() (sebelumnya
tanpa D/E, sudah ditambal) ternyata meluas ke logic.rencana_yudisium_rows()
& logic.wisuda_rows() -- yang muncul di ekspor Excel resmi "Rencana
Yudisium"/"Wisuda" (routes/kelulusan.py) DAN di dokumen SK Yudisium Word
(routes/surat.py::_gen_sk_yudisium).

Latar belakang: baris di rencana_yudisium_rows()/wisuda_rows() dijamin
berasal dari mahasiswa yang sidang.status_kelulusan sudah 'LULUS' (filter
di logic.sync_yudisium_dari_sidang). status_kelulusan itu keputusan
MANUAL tim penguji, independen dari nilai_angka. Kalau nilai_angka rendah
(mis. 30) dikonversi apa adanya lewat nilai_angka_ke_huruf() biasa,
hasilnya "E" -- padahal secara definisi aplikasi sendiri (constants.
NILAI_HURUF_LULUS) huruf E berarti GAGAL/tidak lulus. Baris yang sama, di
dokumen resmi yang sama, menyatakan mahasiswa itu LULUS (itulah premis dia
ada di tabel Yudisium/Wisuda/SK sama sekali) sekaligus "Nilai Huruf: E"
(=tidak lulus) -- kontradiksi tertanam di dokumen resmi.

Diperbaiki dgn constants.nilai_angka_ke_huruf_yudisium(): nilai_angka
mentah tetap tampil apa adanya (transparan), hanya "Nilai Huruf" turunan
di-floor ke "C" (huruf lulus terendah) kalau hasil konversi mentah D/E --
dipakai KHUSUS di rencana_yudisium_rows()/wisuda_rows(), TIDAK menyentuh
nilai_angka_ke_huruf() asli yang masih dipakai apa adanya utk nilai mata
kuliah biasa (nilai.py, semester_pendek.html) di mana D/E memang berarti
tidak lulus sungguhan (tidak ada keputusan panel yang menimpanya).

Tidak diikutkan di paket produksi (murni verifikasi pengembangan).
"""
import os
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

tmpdir = tempfile.mkdtemp()
os.environ["HOME"] = tmpdir

from app import create_app  # noqa: E402
from app import logic as L  # noqa: E402
from app import constants as C  # noqa: E402

FAILS = []


def check(label, cond):
    print(f"[{'OK' if cond else 'FAIL'}] {label}")
    if not cond:
        FAILS.append(label)


db_path = os.path.join(tmpdir, "test.db")
app = create_app(db_path=db_path)
app.config["TESTING"] = True
app.config["WTF_CSRF_ENABLED"] = False
client = app.test_client()
client.post("/login", data={"username": "kaprodi", "password1": "test1234", "password2": "test1234"}, follow_redirects=True)


def buat_mahasiswa_lulus(conn, nim, nama, nilai_angka):
    conn.execute(
        "INSERT INTO mahasiswa(nim, nama, status_ta) VALUES(?,?,?)",
        (nim, nama, "Sudah Sidang"))
    mid = conn.execute("SELECT id FROM mahasiswa WHERE nim=?", (nim,)).fetchone()["id"]
    conn.execute(
        "INSERT INTO sidang(mahasiswa_id, judul_sidang, nilai_angka, status_kelulusan) "
        "VALUES(?,?,?,'LULUS')",
        (mid, f"Skripsi {nama}", nilai_angka))
    conn.commit()
    return mid


with app.app_context():
    conn = app.get_db()

    # -------------------------------------------------------------
    # 1. Unit: fungsi konversi baru sendiri -- D/E di-floor ke C, huruf
    #    lulus lain (A/A-/B+/B/B-/C) TIDAK berubah, kosong tetap kosong.
    # -------------------------------------------------------------
    check("nilai_angka_ke_huruf_yudisium(30) -> 'C' (bukan 'E')",
          C.nilai_angka_ke_huruf_yudisium(30) == "C")
    check("nilai_angka_ke_huruf_yudisium(45) -> 'C' (bukan 'D')",
          C.nilai_angka_ke_huruf_yudisium(45) == "C")
    check("nilai_angka_ke_huruf_yudisium(55) -> 'C' (memang sudah C)",
          C.nilai_angka_ke_huruf_yudisium(55) == "C")
    check("nilai_angka_ke_huruf_yudisium(90) -> 'A' (tidak diubah)",
          C.nilai_angka_ke_huruf_yudisium(90) == "A")
    check("nilai_angka_ke_huruf_yudisium(None) -> '' (tetap kosong)",
          C.nilai_angka_ke_huruf_yudisium(None) == "")
    check("nilai_angka_ke_huruf() ASLI tetap apa adanya (30 -> 'E', tidak disentuh)",
          C.nilai_angka_ke_huruf(30) == "E")

    # -------------------------------------------------------------
    # 2. Skenario persis seperti dilaporkan: mahasiswa LULUS (keputusan
    #    panel) dgn nilai sidang rendah (55) -> Nilai Huruf harus 'C',
    #    BUKAN turunan mentah yg kebetulan sama nilainya di kasus ini,
    #    jadi tambah kasus lebih rendah lagi di bawah (nilai 30 -> 'E' mentah)
    # -------------------------------------------------------------
    buat_mahasiswa_lulus(conn, "2020010001", "Mhs Nilai Rendah 55", 55)
    rows = L.rencana_yudisium_rows(conn)
    row = next(r for r in rows if r["nim"] == "2020010001")
    check("Yudisium: LULUS nilai 55 -> nilai_angka tetap 55 (transparan)", row["nilai_angka"] == 55)
    check("Yudisium: LULUS nilai 55 -> nilai_huruf 'C' (lulus, konsisten)", row["nilai_huruf"] == "C")

    # -------------------------------------------------------------
    # 3. Kasus penting: nilai sidang SANGAT rendah (30) tapi tetap
    #    dinyatakan LULUS oleh panel -- SEBELUM perbaikan ini akan
    #    tercatat "Nilai Huruf: E" di dokumen resmi Yudisium/Wisuda/SK
    #    walau status mahasiswa sudah LULUS. Ini kasus intinya.
    # -------------------------------------------------------------
    buat_mahasiswa_lulus(conn, "2020010002", "Mhs Nilai Sangat Rendah 30", 30)
    rows = L.rencana_yudisium_rows(conn)
    row2 = next(r for r in rows if r["nim"] == "2020010002")
    check("Yudisium: LULUS nilai 30 -> nilai_angka tetap 30 (transparan, tidak dipalsukan)",
          row2["nilai_angka"] == 30)
    check("Yudisium: LULUS nilai 30 -> nilai_huruf 'C' (BUKAN 'E' yg kontradiktif dgn status LULUS)",
          row2["nilai_huruf"] == "C")

    # Set no_sk + status_yudisium=Terlaksana supaya masuk sinkronisasi Wisuda juga
    conn.execute("UPDATE yudisium SET status_yudisium='Terlaksana', no_sk='SK/001/2026' "
                 "WHERE mahasiswa_id=(SELECT id FROM mahasiswa WHERE nim='2020010002')")
    conn.commit()
    wrows = L.wisuda_rows(conn)
    wrow2 = next(r for r in wrows if r["nim"] == "2020010002")
    check("Wisuda: baris yg sama juga sudah 'C' (bukan 'E') via sync dari Yudisium",
          wrow2["nilai_huruf"] == "C")

    # -------------------------------------------------------------
    # 4. Regresi: mahasiswa dgn nilai sidang TINGGI tetap tidak berubah
    #    (memastikan fix ini tidak menyentuh kasus normal).
    # -------------------------------------------------------------
    buat_mahasiswa_lulus(conn, "2020010003", "Mhs Nilai Tinggi 88", 88)
    rows = L.rencana_yudisium_rows(conn)
    row3 = next(r for r in rows if r["nim"] == "2020010003")
    check("Yudisium: LULUS nilai 88 -> nilai_huruf 'A' (tidak terdampak fix)", row3["nilai_huruf"] == "A")

    # -------------------------------------------------------------
    # 5. Regresi: nilai_angka_ke_huruf() ASLI (dipakai di modul Nilai
    #    Semester Pendek, dsb) TIDAK ikut berubah oleh fix ini -- D/E
    #    tetap D/E krn di sana memang tidak ada keputusan panel yg
    #    menimpanya (murni nilai mata kuliah biasa).
    # -------------------------------------------------------------
    check("Modul Nilai biasa: nilai_angka_ke_huruf(30) tetap 'E' (tidak ikut di-floor)",
          C.nilai_angka_ke_huruf(30) == "E")
    check("Modul Nilai biasa: nilai_angka_ke_huruf(45) tetap 'D' (tidak ikut di-floor)",
          C.nilai_angka_ke_huruf(45) == "D")

    # -------------------------------------------------------------
    # 6. Jalur HTTP penuh: ekspor Excel "Rencana Yudisium" -- pastikan
    #    tetap 200 & memuat perbaikan (verifikasi lewat isi file).
    # -------------------------------------------------------------
    r = client.get("/kelulusan/yudisium/ekspor")
    check("GET /kelulusan/yudisium/ekspor -> 200", r.status_code == 200)
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    ws = wb.active
    header = [c.value for c in ws[1]]
    idx_nim = header.index("NIM")
    idx_huruf = header.index("Nilai Huruf")
    ekspor_huruf = None
    for rowcells in ws.iter_rows(min_row=2, values_only=True):
        if rowcells[idx_nim] == "2020010002":
            ekspor_huruf = rowcells[idx_huruf]
            break
    check("Ekspor Excel Rencana Yudisium: baris nilai 30/LULUS -> kolom 'Nilai Huruf' = 'C'",
          ekspor_huruf == "C")

    # -------------------------------------------------------------
    # 7. Jalur dokumen SK Yudisium (Word) -- pastikan ikut memakai
    #    fungsi yg sudah diperbaiki (sama-sama lewat rencana_yudisium_rows).
    # -------------------------------------------------------------
    mid2 = conn.execute("SELECT id FROM mahasiswa WHERE nim='2020010002'").fetchone()["id"]
    r2 = client.post("/surat/buat", data={"mahasiswa_id": mid2, "jenis": "SK Yudisium"})
    check("POST /surat/buat (SK Yudisium) -> 200", r2.status_code == 200)
    import docx
    doc = docx.Document(io.BytesIO(r2.data))
    isi_tabel = []
    for t in doc.tables:
        for trow in t.rows:
            isi_tabel.append([c.text for c in trow.cells])
    ditemukan = any("Nilai Huruf" in cells and "C" in cells for cells in isi_tabel)
    check("Dokumen SK Yudisium (Word): baris 'Nilai Huruf' = 'C' (bukan 'E')", ditemukan)

print("\n=== SELESAI ===")
if FAILS:
    print("ADA YANG GAGAL:", FAILS)
    sys.exit(1)
else:
    print("SEMUA TES LULUS.")
