# -*- coding: utf-8 -*-
"""
import_generic.py — Import Data Generik per-modul (Audit poin 6, diperluas
di Audit Lanjutan poin "Import Modul SDM").

Berbeda dari import_excel.py (migrasi SEKALI JALAN dari 1 workbook lengkap
format lama), modul ini untuk pemakaian RUTIN sehari-hari: admin prodi
mengunduh template kosong 1 sheet per modul, mengisinya, lalu meng-upload
ulang kapan saja — misalnya menambah 40 mahasiswa baru di awal semester,
atau memuat ulang data dosen dari SIMPEG/PDDIKTI.

Diprioritaskan sesuai konfirmasi: Dosen & Mahasiswa (Reguler & RPL) lebih
dulu. Audit Lanjutan menutup celah yang masih tercatat di README ("Import
Excel modul SDM menyusul"): 7 tabel log Modul SDM & Kinerja Dosen
(Pendidikan/Pengajaran, Penelitian, PKM, Penunjang, Luaran, Peran Akademik,
Timeline Karier, Target Kinerja) sebelumnya HANYA bisa diisi satu-satu
lewat form di menu SDM — sekarang semuanya juga bisa diisi lewat
Import Data (template Excel), jadi operator prodi tidak perlu mengetik
ulang manual data yang sudah ada di rekap lama (mis. workbook AKD Excel
Pro atau BKD/SISTER).

Setiap importer dalam IMPORTERS berupa dict:
    label        : nama tampilan
    kategori     : label pengelompokan di dropdown UI ("Data Master"/"SDM & Kinerja Dosen")
    header       : list nama kolom persis urutan di template
    contoh       : 1 baris contoh (list) ditulis di bawah header
    proses_baris : fungsi(conn, row_dict, baris_ke) -> ("tambah"|"update"|"lewati", pesan)

Importer untuk 7 tabel log SDM dibangun oleh SATU factory generik
(`_buat_proses_baris_log_sdm`) alih-alih ditulis berulang 7 kali — sama
persis dengan filosofi TABEL_CONFIG di routes/sdm.py (CRUD-nya juga
ditulis sekali secara generik). Supaya tidak menggandakan data tiap kali
file yang sama diimpor ulang, proses ini idempoten: dicocokkan ke baris
yang SUDAH ADA lewat kombinasi (dosen + field judul/nama utama + tahun),
kalau cocok akan DIPERBARUI, kalau tidak akan DITAMBAHKAN — perilaku yang
sama dengan importer Dosen/Mahasiswa di bawah.
"""

import datetime
import io
import re

import openpyxl
from openpyxl.styles import Font, PatternFill

from app import constants as C
from app import error_utils as EH


def norm(v):
    if v is None:
        return ""
    s = str(v).strip()
    return re.sub(r"\s+", " ", s)


def norm_key(v):
    return norm(v).upper()


def fmt_tanggal(v):
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%d %b %Y")
    return str(v).strip()


# --------------------------------------------------------------------- Dosen
DOSEN_HEADER = [
    "NIDN",
    "Nama*",
    "No HP",
    "Email",
    "NIK",
    "NUPTK",
    "Jabatan Fungsional",
    "Pendidikan Terakhir",
    "Bidang Keahlian",
    "Status Homebase (Homebase/Dosen Luar Prodi/Dosen Luar Fakultas/Dosen Luar PT)",
    "Unit Asal (jika dosen luar)",
    "Prodi Homebase Resmi",
    "No. SK Penugasan",
    "Aktif (Y/N)",
]
DOSEN_CONTOH = [
    "0012345678",
    "Dr. Contoh Nama, M.Kom",
    "081234567890",
    "contoh@kampus.ac.id",
    "3201012345670001",
    "1234567890123456",
    "Lektor",
    "S3",
    "Rekayasa Perangkat Lunak",
    "Homebase",
    "",
    "",
    "",
    "Y",
]


def _dosen_proses(conn, row, baris_ke):
    """Audit Lanjutan — bug ditemukan & dilaporkan: pencocokan baris impor
    ke dosen yang sudah ada punya 2 jalur (NIDN dulu, fallback ke nama
    kalau NIDN di baris kosong), tapi query UPDATE sebelumnya SELALU
    menulis ulang kolom `nidn`/`nip` dgn nilai dari baris impor apa
    adanya — termasuk kalau kosong. Skenario nyata: operator meng-upload
    ulang template hanya utk memperbarui sebagian kolom (mis. email/No
    HP) dan tidak mengisi ulang kolom NIDN di tiap baris (menganggap
    "kalau kosong ya dibiarkan, toh dicocokkan lewat nama") -> NIDN dosen
    yang cocok lewat jalur nama (atau NIP siapa pun yang kolomnya
    dikosongkan, terlepas dari jalur pencocokan) HILANG diam-diam, tanpa
    peringatan apa pun (`hasil['update']` naik normal seperti update
    biasa). Ini serius karena NIDN/NIP adalah kunci identitas resmi
    dosen yang dipakai banyak sistem eksternal (Sister/PDDikti/
    kepegawaian).

    Diperbaiki dgn "jangan timpa dengan kosong" KHUSUS utk kolom
    identitas resmi ini (nidn, nik & nuptk — struktur data SISTER,
    menggantikan NIP) — kalau baris impor kosong utk kolom ini tapi data
    lama sudah ada, nilai lama dipertahankan. Kolom detail lain (No HP,
    Email, Jabatan Fungsional, dst.) SENGAJA tetap memakai perilaku lama
    (ikut nilai baris impor apa adanya, termasuk kalau dikosongkan)
    karena itu memang cara operator mengosongkan kolom tsb secara
    sengaja lewat re-import — hanya identitas resmi yg butuh proteksi
    ekstra karena kalau hilang tidak ada indikasi apa pun ke operator
    bahwa itu terjadi."""
    nama = norm(row.get("Nama*"))
    if not nama:
        return "lewati", f"Baris {baris_ke}: nama kosong, dilewati."
    nidn = norm(row.get("NIDN"))
    nik = norm(row.get("NIK"))
    nuptk = norm(row.get("NUPTK"))
    status_homebase = (
        norm(
            row.get("Status Homebase (Homebase/Dosen Luar Prodi/Dosen Luar Fakultas/Dosen Luar PT)")
        )
        or "Homebase"
    )
    if status_homebase not in C.STATUS_HOMEBASE_LIST:
        status_homebase = "Homebase"
    aktif = 0 if norm(row.get("Aktif (Y/N)")).upper() == "N" else 1
    # Audit Menyeluruh — PHASE 5 (re-check saat implementasi): sejak
    # dosen.status_kepegawaian jadi SATU-SATUNYA sumber kebenaran siklus
    # hidup dosen (lihat routes/dosen.py), `aktif` di sini HARUS diturunkan
    # bersamaan dari kolom yang sama, bukan disetel sendiri-sendiri --
    # kalau tidak, impor massal bisa membuat aktif=0 tapi status_kepegawaian
    # tetap 'Aktif' (atau sebaliknya), dua sumber yang tidak sinkron lagi.
    # Template impor hanya punya "Aktif (Y/N)" (biner), jadi 'N' dipetakan
    # ke 'Nonaktif' generik -- operator bisa mempertajam ke Pindah/Pensiun
    # manual lewat halaman Data Dosen kalau perlu.
    status_kepegawaian = "Aktif" if aktif else "Nonaktif"

    existing = None
    if nidn:
        existing = conn.execute(
            "SELECT id, nidn, nik, nuptk FROM dosen WHERE nidn=?", (nidn,)
        ).fetchone()
    if not existing:
        existing = conn.execute(
            "SELECT id, nidn, nik, nuptk FROM dosen WHERE UPPER(TRIM(nama))=?", (norm_key(nama),)
        ).fetchone()

    nidn_final = nidn or (existing["nidn"] if existing else "")
    nik_final = nik or (existing["nik"] if existing else "")
    nuptk_final = nuptk or (existing["nuptk"] if existing else "")

    nilai = (
        nidn_final,
        nama,
        norm(row.get("No HP")),
        norm(row.get("Email")),
        aktif,
        status_kepegawaian,
        nik_final,
        nuptk_final,
        norm(row.get("Jabatan Fungsional")),
        norm(row.get("Pendidikan Terakhir")),
        norm(row.get("Bidang Keahlian")),
        status_homebase,
        norm(row.get("Unit Asal (jika dosen luar)")),
        norm(row.get("Prodi Homebase Resmi")),
        norm(row.get("No. SK Penugasan")),
    )
    if existing:
        conn.execute(
            "UPDATE dosen SET nidn=?, nama=?, no_hp=?, email=?, aktif=?, status_kepegawaian=?, "
            "nik=?, nuptk=?, jabatan_fungsional=?, pendidikan_terakhir=?, bidang_keahlian=?, "
            "status_homebase=?, unit_asal=?, prodi_homebase=?, sk_penugasan=? WHERE id=?",
            nilai + (existing["id"],),
        )
        return "update", f"Baris {baris_ke}: {nama} — diperbarui."
    conn.execute(
        "INSERT INTO dosen(nidn,nama,no_hp,email,aktif,status_kepegawaian,nik,nuptk,jabatan_fungsional,"
        "pendidikan_terakhir,bidang_keahlian,status_homebase,unit_asal,prodi_homebase,sk_penugasan) "
        "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        nilai,
    )
    return "tambah", f"Baris {baris_ke}: {nama} — ditambahkan."


# ---------------------------------------------------------------- Mahasiswa
MAHASISWA_HEADER = [
    "NIM*",
    "Nama*",
    "Jenis Kelamin (L/P)",
    "Tempat Lahir",
    "Tanggal Lahir",
    "No HP",
    "Email/NIK",
    "Angkatan",
    "Status (Aktif/Cuti/Non-Aktif/Drop Out)",
    "Skema (Reguler/RPL)",
]
MAHASISWA_CONTOH = [
    "2023010001",
    "Contoh Mahasiswa",
    "L",
    "Kota Contoh",
    "01/01/2000",
    "081234567890",
    "contoh@student.ac.id",
    "2023",
    "Aktif",
    "Reguler",
]


def _mahasiswa_proses(conn, row, baris_ke):
    nim = norm(row.get("NIM*"))
    nama = norm(row.get("Nama*"))
    if not nim or not nama:
        return "lewati", f"Baris {baris_ke}: NIM/Nama kosong, dilewati."
    status = norm(row.get("Status (Aktif/Cuti/Non-Aktif/Drop Out)")) or "Aktif"
    if status not in C.STATUS_MHS_LIST:
        status = "Aktif"
    skema = norm(row.get("Skema (Reguler/RPL)")) or "Reguler"
    if skema not in ("Reguler", "RPL"):
        skema = "Reguler"
    jk = norm(row.get("Jenis Kelamin (L/P)"))
    tempat_lahir = norm(row.get("Tempat Lahir"))
    tgl_lahir = fmt_tanggal(row.get("Tanggal Lahir"))
    no_hp = norm(row.get("No HP"))
    email_nik = norm(row.get("Email/NIK"))
    angkatan = norm(row.get("Angkatan")) or (nim[:4] if len(nim) >= 4 and nim[:4].isdigit() else "")

    existing = conn.execute("SELECT id FROM mahasiswa WHERE nim=?", (nim,)).fetchone()
    if existing:
        conn.execute(
            "UPDATE mahasiswa SET nama=?, jk=?, tempat_lahir=?, tgl_lahir=?, no_hp=?, "
            "email_nik=?, angkatan=?, status=?, skema=? WHERE id=?",
            (
                nama,
                jk,
                tempat_lahir,
                tgl_lahir,
                no_hp,
                email_nik,
                angkatan,
                status,
                skema,
                existing["id"],
            ),
        )
        return "update", f"Baris {baris_ke}: {nim} — {nama} diperbarui."
    conn.execute(
        "INSERT INTO mahasiswa(nim,nama,jk,tempat_lahir,tgl_lahir,no_hp,email_nik,angkatan,"
        "status,status_ta,skema) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
        (
            nim,
            nama,
            jk,
            tempat_lahir,
            tgl_lahir,
            no_hp,
            email_nik,
            angkatan,
            status,
            C.STATUS_TA_BELUM,
            skema,
        ),
    )
    return "tambah", f"Baris {baris_ke}: {nim} — {nama} ditambahkan."


# ----------------------------------------------------- Modul SDM & Kinerja Dosen
# 7 tabel log per-dosen (lihat routes/sdm.py TABEL_CONFIG) dibangun lewat 1
# factory generik. Setiap importer butuh kolom "NIDN atau Nama Dosen*" utk
# mencocokkan ke dosen yang SUDAH ADA di database (dosen wajib diimpor/
# ditambah lebih dulu lewat importer "dosen" di atas — SDM tidak membuat
# dosen baru secara implisit, supaya tidak menggandakan data dosen dari 2
# jalur berbeda).

KOL_DOSEN = "NIDN atau Nama Dosen*"
KOL_TAHUN_AJARAN = "Tahun Ajaran (mis. 2025/2026, boleh kosong)"
KOL_SEMESTER = "Semester (Ganjil/Genap/Antara, boleh kosong)"


def _cari_dosen_id(conn, nidn_atau_nama):
    """Cocokkan ke dosen yang sudah ada di database — via NIDN dulu (lebih
    presisi), baru nama (case-insensitive, trim spasi ganda). Tidak pernah
    membuat baris dosen baru dari sini."""
    key = norm(nidn_atau_nama)
    if not key:
        return None
    row = conn.execute("SELECT id FROM dosen WHERE nidn=? AND nidn!=''", (key,)).fetchone()
    if row:
        return row["id"]
    row = conn.execute(
        "SELECT id FROM dosen WHERE UPPER(TRIM(nama))=?", (norm_key(nidn_atau_nama),)
    ).fetchone()
    return row["id"] if row else None


def _generate_kode_sdm(conn, table, prefix, tahun_akademik):
    """Format sama dgn routes/sdm.py `_generate_kode`: PEN-2026-003 (tahun
    dari tahun_akademik, urut per-prefix+tahun)."""
    tahun = "XXXX"
    if tahun_akademik:
        tahun = str(tahun_akademik).split("/")[0].strip() or "XXXX"
    n = conn.execute(
        f"SELECT COUNT(*) FROM {table} WHERE kode LIKE ?", (f"{prefix}-{tahun}-%",)
    ).fetchone()[0]
    return f"{prefix}-{tahun}-{n + 1:03d}"


def _buat_proses_baris_log_sdm(
    table,
    wajib_label,
    wajib_col,
    kolom_map,
    kode_prefix=None,
    pakai_periode=True,
    kunci_tambahan=None,
):
    """Bangun fungsi proses_baris() untuk 1 tabel log SDM.

    kolom_map      : list (label_header, nama_kolom_db, tipe) — tipe salah
                      satu dari "text"/"number"/"int".
    wajib_label     : label header kolom yang wajib diisi (dipakai jadi
                      penanda identitas baris di pesan hasil import).
    wajib_col       : nama kolom DB yang berpasangan dengan wajib_label —
                      dipakai bareng dosen_id (& kunci_tambahan) utk
                      mendeteksi baris yang SUDAH ADA (supaya idempoten,
                      tidak menggandakan data kalau file yang sama
                      diimpor ulang).
    kode_prefix     : kalau diisi, kolom `kode` auto-generate (mis. PEN/PKM/LUR).
    pakai_periode   : kalau True, sertakan kolom Tahun Ajaran/Semester dan
                      simpan sbg tahun_akademik/semester (dipakai filter &
                      Target Kinerja on-the-fly di menu SDM).
    kunci_tambahan  : nama kolom DB tambahan (selain dosen_id + wajib_col)
                      utk pencocokan baris yang sudah ada, mis. ["tahun"]
                      utk Target Kinerja (kategori sama boleh beda tahun).
    """
    kunci_tambahan = kunci_tambahan or ([] if not pakai_periode else ["tahun_akademik"])

    def _proses(conn, row, baris_ke):
        dosen_id = _cari_dosen_id(conn, row.get(KOL_DOSEN))
        if not dosen_id:
            return "lewati", (
                f"Baris {baris_ke}: dosen '{norm(row.get(KOL_DOSEN))}' tidak ditemukan di "
                "database (cocokkan via NIDN/nama) — impor/tambah data dosen dulu, dilewati."
            )
        wajib_val = norm(row.get(wajib_label))
        if not wajib_val:
            return "lewati", f"Baris {baris_ke}: kolom wajib '{wajib_label}' kosong, dilewati."

        val_by_col = {}
        for label, col, tipe in kolom_map:
            raw = row.get(label)
            if tipe == "number":
                try:
                    v = float(raw) if raw not in (None, "") else 0
                except (TypeError, ValueError):
                    v = 0
            elif tipe == "int":
                try:
                    v = int(float(raw)) if raw not in (None, "") else None
                except (TypeError, ValueError):
                    v = None
                if v is None and "*" in label:
                    return (
                        "lewati",
                        f"Baris {baris_ke}: kolom wajib '{label}' kosong/bukan angka, dilewati.",
                    )
            else:
                v = norm(raw)
            val_by_col[col] = v

        if pakai_periode:
            val_by_col["tahun_akademik"] = norm(row.get(KOL_TAHUN_AJARAN))
            # Semester disimpan apa adanya (kolom TEXT bebas, dipakai cache
            # tampilan) — tidak dipaksa cocok C.SEMESTER_LIST supaya import
            # tetap jalan walau operator mengisi variasi penulisan.
            val_by_col["semester"] = norm(row.get(KOL_SEMESTER))

        cols = list(val_by_col.keys())
        vals = [val_by_col[c] for c in cols]

        where_parts = ["dosen_id=?"]
        where_vals = [dosen_id]
        for kcol in kunci_tambahan:
            where_parts.append(f"{kcol}=?")
            where_vals.append(val_by_col.get(kcol, ""))
        existing = conn.execute(
            f"SELECT id FROM {table} WHERE {' AND '.join(where_parts)} AND {wajib_col}=?",
            (*where_vals, wajib_val),
        ).fetchone()

        if existing:
            set_clause = ", ".join(f"{c}=?" for c in cols)
            conn.execute(f"UPDATE {table} SET {set_clause} WHERE id=?", (*vals, existing["id"]))
            return "update", f"Baris {baris_ke}: {wajib_val} — diperbarui."

        insert_cols = ["dosen_id"] + cols
        insert_vals = [dosen_id] + vals
        if kode_prefix:
            kode = _generate_kode_sdm(
                conn, table, kode_prefix, val_by_col.get("tahun_akademik", "")
            )
            insert_cols = ["kode"] + insert_cols
            insert_vals = [kode] + insert_vals
        placeholders = ", ".join("?" for _ in insert_vals)
        conn.execute(
            f"INSERT INTO {table}({', '.join(insert_cols)}) VALUES({placeholders})", insert_vals
        )
        return "tambah", f"Baris {baris_ke}: {wajib_val} — ditambahkan."

    return _proses


def _header_sdm(kolom_map, pakai_periode=True):
    header = [KOL_DOSEN]
    if pakai_periode:
        header += [KOL_TAHUN_AJARAN, KOL_SEMESTER]
    header += [label for label, _col, _tipe in kolom_map]
    return header


def _contoh_sdm(nidn_contoh, isian, pakai_periode=True, tahun="2025/2026", semester="Ganjil"):
    contoh = [nidn_contoh]
    if pakai_periode:
        contoh += [tahun, semester]
    contoh += isian
    return contoh


# --- Pendidikan & Pengajaran ---
_PENDIDIKAN_KOLOM = [
    ("Mata Kuliah*", "mata_kuliah", "text"),
    ("Kode MK", "kode_mk", "text"),
    ("SKS", "sks", "text"),
    ("Jumlah Kelas", "jumlah_kelas", "text"),
    ("Jumlah Mahasiswa", "jumlah_mahasiswa", "text"),
    ("Peran", "peran", "text"),
    (f"Status ({'/'.join(C.STATUS_AKTIVITAS_SDM_LIST)})", "status", "text"),
    ("Catatan", "catatan", "text"),
]

# --- Penelitian ---
_PENELITIAN_KOLOM = [
    ("Judul*", "judul", "text"),
    (f"Skema ({'/'.join(C.SKEMA_PENELITIAN_PKM_LIST)})", "skema", "text"),
    (f"Sumber Dana ({'/'.join(C.SUMBER_DANA_LIST)})", "sumber_dana", "text"),
    ("Nominal", "nominal", "number"),
    ("Pelaksana", "pelaksana", "text"),
    ("Tgl Publish", "tgl_publish", "text"),
    ("Jurnal", "jurnal", "text"),
    ("Jilid", "jilid", "text"),
    ("Vol", "volume", "text"),
    ("Halaman", "halaman", "text"),
    (f"Status ({'/'.join(C.STATUS_AKTIVITAS_SDM_LIST)})", "status", "text"),
    (f"Jenis Luaran ({'/'.join(C.JENIS_LUARAN_LIST)})", "jenis_luaran", "text"),
    ("DOI", "doi", "text"),
    ("ISSN/ISBN", "issn_isbn", "text"),
    ("URL", "url", "text"),
    ("Lokasi Bukti", "lokasi_bukti", "text"),
    ("Catatan", "catatan", "text"),
]

# --- Pengabdian kepada Masyarakat (PKM) ---
_PKM_KOLOM = [
    ("Judul*", "judul", "text"),
    ("Jenis", "jenis", "text"),
    (f"Skema ({'/'.join(C.SKEMA_PENELITIAN_PKM_LIST)})", "skema", "text"),
    ("Lokasi", "lokasi", "text"),
    ("Mitra", "mitra", "text"),
    ("Dana", "dana", "number"),
    (f"Status ({'/'.join(C.STATUS_AKTIVITAS_SDM_LIST)})", "status", "text"),
    (f"Jenis Luaran ({'/'.join(C.JENIS_LUARAN_LIST)})", "jenis_luaran", "text"),
    ("URL", "url", "text"),
    ("Lokasi Bukti", "lokasi_bukti", "text"),
    ("Catatan", "catatan", "text"),
]

# --- Penunjang ---
_PENUNJANG_KOLOM = [
    ("Jenis Penunjang", "jenis_penunjang", "text"),
    ("Nama Kegiatan/Instansi*", "nama_kegiatan", "text"),
    ("Peran", "peran", "text"),
    ("Tanggal", "tanggal", "text"),
    (f"Status ({'/'.join(C.STATUS_AKTIVITAS_SDM_LIST)})", "status", "text"),
    ("URL", "url", "text"),
    ("Lokasi Bukti", "lokasi_bukti", "text"),
    ("Catatan", "catatan", "text"),
]

# --- Luaran (Publikasi/HKI/Buku/Prosiding/Seminar/Sertifikat/Penghargaan) ---
_LUARAN_KOLOM = [
    (f"Jenis Luaran* ({'/'.join(C.JENIS_LUARAN_LIST)})", "jenis_luaran", "text"),
    ("Judul/Nama*", "judul", "text"),
    ("Penulis/Pihak Terkait", "penulis_terkait", "text"),
    ("Nomor Identitas (DOI/ISSN/ISBN/No.HKI/No.Sertifikat)", "nomor_identitas", "text"),
    ("Penerbit/Instansi", "penerbit_instansi", "text"),
    (f"Sumber Dana ({'/'.join(C.SUMBER_DANA_LIST)})", "sumber_dana", "text"),
    (f"Status ({'/'.join(C.STATUS_AKTIVITAS_SDM_LIST)})", "status", "text"),
    ("Masa Berlaku (khusus Sertifikat, format YYYY-MM-DD)", "masa_berlaku", "text"),
    ("URL", "url", "text"),
    ("Lokasi Bukti", "lokasi_bukti", "text"),
    ("Keterangan Tambahan", "keterangan_tambahan", "text"),
    ("Catatan", "catatan", "text"),
]

# --- Peran Akademik ---
_PERAN_AKADEMIK_KOLOM = [
    (f"Jenis Peran* ({'/'.join(C.JENIS_PERAN_AKADEMIK_LIST)})", "jenis_peran", "text"),
    ("Nama Instansi/Kegiatan*", "nama_instansi_kegiatan", "text"),
    ("Peran/Jabatan", "peran_jabatan", "text"),
    ("Tgl Mulai", "tgl_mulai", "text"),
    ("Tgl Selesai / Masa Berlaku", "tgl_selesai", "text"),
    (f"Status ({'/'.join(C.STATUS_AKTIVITAS_SDM_LIST)})", "status", "text"),
    ("URL", "url", "text"),
    ("Lokasi Bukti", "lokasi_bukti", "text"),
    ("Catatan", "catatan", "text"),
]

# --- Timeline Karier (tidak terikat periode akademik, pakai tanggal sendiri) ---
_TIMELINE_KOLOM = [
    (f"Jenis Perubahan* ({'/'.join(C.JENIS_PERUBAHAN_KARIER_LIST)})", "jenis_perubahan", "text"),
    ("Keterangan/Nama*", "keterangan", "text"),
    ("Nomor SK", "no_sk", "text"),
    ("Tgl Mulai", "tgl_mulai", "text"),
    ("Tgl Berakhir/Target Berikutnya", "tgl_berakhir_target", "text"),
    ("Instansi Penerbit", "instansi_penerbit", "text"),
    (f"Status ({'/'.join(C.STATUS_AKTIVITAS_SDM_LIST)})", "status", "text"),
    ("Lokasi Bukti", "lokasi_bukti", "text"),
    ("Catatan", "catatan", "text"),
]

# --- Target Kinerja Tahunan (tidak terikat periode akademik, pakai tahun angka) ---
_KOL_KATEGORI_TARGET = f"Kategori* ({'/'.join(C.KATEGORI_TARGET_KINERJA_LIST)})"
_TARGET_KOLOM = [
    ("Tahun*", "tahun", "int"),
    (_KOL_KATEGORI_TARGET, "kategori", "text"),
    ("Target", "target_angka", "number"),
    ("Keterangan", "keterangan", "text"),
]


IMPORTERS = {
    "dosen": {
        "label": "Data Dosen",
        "kategori": "Data Master",
        "header": DOSEN_HEADER,
        "contoh": DOSEN_CONTOH,
        "proses_baris": _dosen_proses,
    },
    "mahasiswa": {
        "label": "Data Mahasiswa (Reguler & RPL)",
        "kategori": "Data Master",
        "header": MAHASISWA_HEADER,
        "contoh": MAHASISWA_CONTOH,
        "proses_baris": _mahasiswa_proses,
    },
    "sdm_pendidikan": {
        "label": "SDM — Pendidikan & Pengajaran",
        "kategori": "SDM & Kinerja Dosen",
        "header": _header_sdm(_PENDIDIKAN_KOLOM),
        "contoh": _contoh_sdm(
            "0012345678",
            [
                "Rekayasa Perangkat Lunak",
                "IF301",
                "3",
                "2",
                "60",
                "Pengampu",
                "Selesai",
                "",
            ],
        ),
        "proses_baris": _buat_proses_baris_log_sdm(
            "aktivitas_pendidikan",
            "Mata Kuliah*",
            "mata_kuliah",
            _PENDIDIKAN_KOLOM,
        ),
    },
    "sdm_penelitian": {
        "label": "SDM — Penelitian",
        "kategori": "SDM & Kinerja Dosen",
        "header": _header_sdm(_PENELITIAN_KOLOM),
        "contoh": _contoh_sdm(
            "0012345678",
            [
                "Contoh Judul Penelitian",
                "Penelitian Dasar",
                "DIKTI/DIKTISAINTEK",
                "15000000",
                "Ketua",
                "01 Jun 2025",
                "Jurnal Contoh",
                "12",
                "3",
                "1-10",
                "Published",
                "Publikasi",
                "10.1234/contoh",
                "1234-5678",
                "https://contoh.ac.id",
                "",
                "",
            ],
        ),
        "proses_baris": _buat_proses_baris_log_sdm(
            "aktivitas_penelitian",
            "Judul*",
            "judul",
            _PENELITIAN_KOLOM,
            kode_prefix="PEN",
        ),
    },
    "sdm_pkm": {
        "label": "SDM — Pengabdian Masyarakat (PKM)",
        "kategori": "SDM & Kinerja Dosen",
        "header": _header_sdm(_PKM_KOLOM),
        "contoh": _contoh_sdm(
            "0012345678",
            [
                "Contoh Judul PKM",
                "Pelatihan",
                "PKM Reguler",
                "Desa Contoh",
                "Mitra Contoh",
                "5000000",
                "Selesai",
                "Publikasi",
                "https://contoh.ac.id",
                "",
                "",
            ],
        ),
        "proses_baris": _buat_proses_baris_log_sdm(
            "aktivitas_pkm",
            "Judul*",
            "judul",
            _PKM_KOLOM,
            kode_prefix="PKM",
        ),
    },
    "sdm_penunjang": {
        "label": "SDM — Penunjang",
        "kategori": "SDM & Kinerja Dosen",
        "header": _header_sdm(_PENUNJANG_KOLOM),
        "contoh": _contoh_sdm(
            "0012345678",
            [
                "Kepanitiaan",
                "Panitia Wisuda",
                "Anggota",
                "01 Jun 2025",
                "Selesai",
                "",
                "",
                "",
            ],
        ),
        "proses_baris": _buat_proses_baris_log_sdm(
            "aktivitas_penunjang",
            "Nama Kegiatan/Instansi*",
            "nama_kegiatan",
            _PENUNJANG_KOLOM,
        ),
    },
    "sdm_luaran": {
        "label": "SDM — Luaran (Publikasi/HKI/Buku/dst)",
        "kategori": "SDM & Kinerja Dosen",
        "header": _header_sdm(_LUARAN_KOLOM),
        "contoh": _contoh_sdm(
            "0012345678",
            [
                "Publikasi",
                "Contoh Judul Artikel",
                "Nama Penulis",
                "1234-5678",
                "Nama Jurnal/Penerbit",
                "Mandiri",
                "Published",
                "",
                "https://contoh.ac.id",
                "",
                "",
                "",
            ],
        ),
        "proses_baris": _buat_proses_baris_log_sdm(
            "luaran_dosen",
            "Judul/Nama*",
            "judul",
            _LUARAN_KOLOM,
            kode_prefix="LUR",
        ),
    },
    "sdm_peran_akademik": {
        "label": "SDM — Peran Akademik",
        "kategori": "SDM & Kinerja Dosen",
        "header": _header_sdm(_PERAN_AKADEMIK_KOLOM),
        "contoh": _contoh_sdm(
            "0012345678",
            [
                "Reviewer",
                "Contoh Jurnal/Kegiatan",
                "Reviewer",
                "01 Jan 2025",
                "31 Des 2025",
                "Selesai",
                "https://contoh.ac.id",
                "",
                "",
            ],
        ),
        "proses_baris": _buat_proses_baris_log_sdm(
            "peran_akademik_dosen",
            "Nama Instansi/Kegiatan*",
            "nama_instansi_kegiatan",
            _PERAN_AKADEMIK_KOLOM,
        ),
    },
    "sdm_timeline": {
        "label": "SDM — Timeline Karier",
        "kategori": "SDM & Kinerja Dosen",
        "header": _header_sdm(_TIMELINE_KOLOM, pakai_periode=False),
        "contoh": _contoh_sdm(
            "0012345678",
            [
                "Jabatan Fungsional",
                "Lektor",
                "SK.123/2025",
                "01 Jan 2025",
                "",
                "Kemdiktisaintek",
                "Selesai",
                "",
            ],
            pakai_periode=False,
        ),
        "proses_baris": _buat_proses_baris_log_sdm(
            "timeline_karier_dosen",
            "Keterangan/Nama*",
            "keterangan",
            _TIMELINE_KOLOM,
            pakai_periode=False,
            kunci_tambahan=["jenis_perubahan", "tgl_mulai"],
        ),
    },
    "sdm_target": {
        "label": "SDM — Target Kinerja Tahunan",
        "kategori": "SDM & Kinerja Dosen",
        "header": _header_sdm(_TARGET_KOLOM, pakai_periode=False),
        "contoh": _contoh_sdm(
            "0012345678", ["2025", "Publikasi", "2", "Target tahunan"], pakai_periode=False
        ),
        "proses_baris": _buat_proses_baris_log_sdm(
            "target_kinerja_dosen",
            _KOL_KATEGORI_TARGET,
            "kategori",
            _TARGET_KOLOM,
            pakai_periode=False,
            kunci_tambahan=["tahun"],
        ),
    },
}


INVALID_SHEET_CHARS = set("[]:*?/\\")


def buat_template(modul):
    """Bangun workbook template kosong (header + 1 baris contoh bertanda
    abu-abu) untuk diunduh & diisi user. Mengembalikan BytesIO siap kirim."""
    info = IMPORTERS[modul]
    wb = openpyxl.Workbook()
    ws = wb.active
    # Nama sheet Excel tidak boleh mengandung karakter [ ] : * ? / \ dan
    # maksimal 31 karakter — beberapa label modul memakai "/" (mis.
    # "Data Mahasiswa (Reguler & RPL)" aman, tapi "SDM — ... (Publikasi/HKI/dst)"
    # tidak), jadi disaring dulu supaya buat_template() tidak error.
    judul_sheet = "".join(ch for ch in info["label"] if ch not in INVALID_SHEET_CHARS)
    ws.title = (judul_sheet or "Template")[:31]
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    for col, judul in enumerate(info["header"], start=1):
        cell = ws.cell(row=1, column=col, value=judul)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    contoh_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    for col, val in enumerate(info["contoh"], start=1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.fill = contoh_fill
        cell.font = Font(italic=True, color="6B7280")
    for col in range(1, len(info["header"]) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 24
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def proses_upload(modul, file_stream, conn):
    """Baca file .xlsx yang diupload, jalankan proses_baris per baris
    (mulai baris 2 — baris contoh di template dianggap baris data biasa
    kalau user lupa menghapusnya & isinya bukan lagi 'Contoh...', jadi
    dilewati otomatis lewat validasi kunci wajib per importer).
    Mengembalikan dict ringkasan: {tambah, update, lewati, detail: [...]}."""
    info = IMPORTERS[modul]
    # Audit Lanjutan (Import Excel) — temuan: berbeda dari import_excel.py
    # (run_import(), dipanggil di routes/pengaturan.py:import_export()
    # yang MEMANG dibungkus try/except), pemanggil fungsi ini
    # (routes/pengaturan.py:import_generik_proses()) TIDAK membungkus
    # panggilan ke proses_upload() sama sekali. Validasi upload sebelumnya
    # hanya memeriksa EKSTENSI nama file (".xlsx") — file apa pun yang
    # namanya diganti jadi *.xlsx (mis. .txt/.csv yang di-rename, atau
    # file .xlsx yang korup) akan lolos pemeriksaan ekstensi lalu meledak
    # di openpyxl.load_workbook() sebagai galat mentah (BadZipFile/
    # KeyError/InvalidFileException) -> 500 Internal Server Error,
    # bukan pesan ramah seperti pola error_utils.py di seluruh aplikasi.
    # Dipertahankan di titik ini (bukan di route) supaya SEMUA pemanggil
    # proses_upload() di masa depan otomatis ikut terlindungi.
    try:
        wb = openpyxl.load_workbook(file_stream, data_only=True)
    except Exception as e:
        if EH.pesan_ramah_db(e) is None:
            EH.logger.exception("Import Generik (%s) — file .xlsx tidak bisa dibuka", modul)
        return {
            "error": (
                "File tidak bisa dibuka sebagai workbook Excel — pastikan file "
                "benar-benar hasil isian template .xlsx (bukan file lain yang "
                "sekadar diganti namanya) dan tidak korup."
            )
        }
    ws = wb.active
    header_actual = [norm(c.value) for c in ws[1]]
    if header_actual[: len(info["header"])] != info["header"]:
        return {
            "error": (
                "Header file tidak cocok dengan template. Unduh ulang template "
                f"terbaru untuk modul '{info['label']}' dan jangan mengubah urutan/nama kolom."
            )
        }
    hasil = {"tambah": 0, "update": 0, "lewati": 0, "detail": []}
    for r in range(2, ws.max_row + 1):
        row_vals = [c.value for c in ws[r]]
        if all(v in (None, "") for v in row_vals):
            continue
        row = dict(zip(info["header"], row_vals))
        try:
            status, pesan = info["proses_baris"](conn, row, r)
        except Exception as e:
            # Audit temuan sistemik — sama seperti import_excel.py: pesan
            # ke operator diterjemahkan ke bahasa manusia, sementara galat
            # yang tidak dikenali (kemungkinan bug) tetap dicatat ke log.
            if EH.pesan_ramah_db(e) is None:
                EH.logger.exception("Import Generik (%s) — gagal proses baris %s", modul, r)
            status, pesan = "lewati", f"Baris {r}: {EH.pesan_ramah_import(e)}"
        hasil[status if status in ("tambah", "update") else "lewati"] += 1
        hasil["detail"].append(pesan)
    conn.commit()
    return hasil
