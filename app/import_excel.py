# -*- coding: utf-8 -*-
"""
import_excel.py — Migrasi data dari workbook Excel asal ke database SQLite.

Versi web: logika inti sama persis dengan aplikasi desktop (lihat catatan
audit di bawah), dibungkus jadi fungsi run_import() yang dipanggil dari
route Flask, dan mengembalikan ringkasan (bukan print ke konsol).

Perbaikan dibanding versi desktop asal:
  - Pencocokan dosen saat impor sekarang juga dicek ke DATABASE (bukan cuma
    dalam 1 kali proses import), supaya meng-impor file yang sama 2x tidak
    lagi menggandakan baris dosen.

Catatan audit yang ditangani saat migrasi (lihat AUDIT_LAPORAN.md):
  - Taksonomi status (kolom K 'Status Pengajuan TA') disamakan ke satu kamus
    resmi (constants.STATUS_TA_LIST): "Lulus"->"LULUS", dst.
  - Baris pembatas tahap ("▶ TAHAP 1 ...") dan baris kosong dilewati otomatis.
  - Pencocokan nama dosen dilakukan case-insensitive & trim spasi ganda,
    supaya inkonsistensi ejaan seperti kasus "HALIMATUSSADIAH" vs
    "HALIMATUS SADIAH" pada file asal tidak menggandakan data dosen.
"""

import datetime
import re
import sqlite3
import sys

import openpyxl

from app import constants as C
from app import db
from app import error_utils as EH

STATUS_TA_MAP = {
    "belum mengajukan judul": C.STATUS_TA_BELUM,
    "mengajukan judul": C.STATUS_TA_MENGAJUKAN,
    "proses bimbingan": C.STATUS_TA_BIMBINGAN,
    "sudah sidang": C.STATUS_TA_SUDAH_SIDANG,
    "lulus": C.STATUS_TA_LULUS,
    "tidak lulus": C.STATUS_TA_TIDAK_LULUS,
    "tunda": C.STATUS_TA_TUNDA,
    "menunggu wisuda": C.STATUS_TA_MENUNGGU_WISUDA,
}

# Audit Menyeluruh — P0 #10 (bagian impor Excel). Sebelum perbaikan ini,
# `seminar.status` dan `sidang.status_kelulusan` diisi LANGSUNG dari teks
# sel Excel (lewat norm(), yang cuma trim spasi) tanpa dicocokkan ke
# C.STATUS_SEMINAR_LIST / C.STATUS_KELULUSAN_SIDANG sama sekali — beda
# dari status_ta di atas yang SUDAH lebih dulu punya STATUS_TA_MAP.
# File Excel lama sering tidak konsisten kapitalisasinya ("Selesai" vs
# "SELESAI", "Lulus" vs "LULUS"), dan nilai yang tidak cocok persis akan
# lolos tersimpan tapi diam-diam tidak pernah terhitung status_sidang_
# mahasiswa()/status_seminar_mahasiswa() (exact-match), yang pada
# gilirannya membuat recalculate_status_ta() menganggap mahasiswa itu
# belum sidang/seminar sama sekali walau datanya sebenarnya ada.
STATUS_SEMINAR_MAP = {s.lower(): s for s in C.STATUS_SEMINAR_LIST}
STATUS_SIDANG_MAP = {s.lower(): s for s in C.STATUS_KELULUSAN_SIDANG}


def _normalisasi_status_seminar(raw):
    return STATUS_SEMINAR_MAP.get(norm(raw).lower(), "Terdaftar")


def _normalisasi_status_sidang(raw):
    """Beda dari _normalisasi_status_seminar() — TIDAK ada fallback default
    di sini (kembalikan None kalau tidak cocok whitelist apa pun), karena
    status_kelulusan yang salah tafsir sebagai default tertentu jauh lebih
    berbahaya (bisa keliru dianggap LULUS/TIDAK LULUS) daripada dibiarkan
    NULL dan diminta operator melengkapi manual lewat halaman Sidang."""
    return STATUS_SIDANG_MAP.get(norm(raw).lower())


def norm(s):
    if s is None:
        return ""
    s = str(s).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def norm_key(s):
    return norm(s).upper()


def fmt(v):
    """Format tanggal/waktu Excel jadi teks yang stabil untuk SQLite."""
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%d %b %Y")
    if isinstance(v, datetime.time):
        return v.strftime("%H:%M")
    return str(v).strip()


def run_import(xlsx_path, conn, skema_default="Reguler"):
    """Jalankan migrasi dari file .xlsx ke koneksi DB yang sudah dibuka.
    Mengembalikan list baris ringkasan (string) untuk ditampilkan ke user.
    Aman dijalankan berulang kali (idempoten) untuk semua sheet."""
    ringkasan = []
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # 1) Pengaturan (dari sheet Panduan, blok parameter baris 48-59)
    if "📖 Panduan" in wb.sheetnames:
        wp = wb["📖 Panduan"]
        pairs = [
            ("tahun_akademik_aktif", "B48"),
            ("nama_tahap_1", "B49"),
            ("nama_tahap_2", "B50"),
            ("nama_institusi", "B51"),
            ("nama_prodi", "B52"),
            ("nama_fakultas", "B53"),
            ("tarif_honor_seminar", "B54"),
            ("tarif_honor_penguji_sidang", "B55"),
            ("tarif_honor_pembimbing_1", "B57"),
            ("tarif_honor_pembimbing_2", "B58"),
            ("ambang_beban_dosen", "B59"),
        ]
        for key, cell in pairs:
            v = wp[cell].value
            if v is not None and str(v).strip() != "":
                db.set_setting(conn, key, str(v))
        ringkasan.append("Pengaturan / parameter tersalin dari sheet Panduan")

        # Fase Fondasi (Audit poin 1 & 2): sinkronkan juga ke struktur baru
        # tahun_ajaran/periode_akademik/tahap_pengajuan, supaya migrasi dari
        # workbook lama langsung siap dipakai wizard "Buka Tahun Ajaran" &
        # dropdown semester terkunci di modul-modul yang sudah dimigrasikan.
        kode_ta = db.get_setting(conn, "tahun_akademik_aktif", "").strip()
        if kode_ta:
            _, periode_ids = db.buka_tahun_ajaran(conn, kode_ta, aktifkan="Ganjil")
            ganjil_id = periode_ids.get("Ganjil")
            ada_tahap = conn.execute(
                "SELECT COUNT(*) c FROM tahap_pengajuan WHERE periode_akademik_id=?",
                (ganjil_id,),
            ).fetchone()["c"]
            if ganjil_id and not ada_tahap:
                for urutan, key in ((1, "nama_tahap_1"), (2, "nama_tahap_2")):
                    nama_t = db.get_setting(conn, key, "").strip()
                    if nama_t:
                        conn.execute(
                            "INSERT INTO tahap_pengajuan(periode_akademik_id, urutan, nama) VALUES(?,?,?)",
                            (ganjil_id, urutan, nama_t),
                        )
                conn.commit()
            ringkasan.append(
                f"Tahun ajaran {kode_ta} disinkronkan ke struktur Ganjil/Genap/Antara baru"
            )

    # 2) Dosen — dicocokkan ke DATABASE dulu (bukan cuma dalam 1x proses ini)
    dosen_id_by_name = {}
    for row in conn.execute("SELECT id, nama FROM dosen"):
        dosen_id_by_name[norm_key(row["nama"])] = row["id"]
    n_dosen_baru = 0
    n_dosen_nidn_bentrok = 0
    if "Data Dosen" in wb.sheetnames:
        ws = wb["Data Dosen"]
        for r in range(7, ws.max_row + 1):
            nama = norm(ws.cell(row=r, column=3).value)
            if not nama:
                continue
            key = norm_key(nama)
            if key in dosen_id_by_name:
                continue
            nidn = norm(ws.cell(row=r, column=2).value)
            no_hp = norm(ws.cell(row=r, column=5).value)
            email = norm(ws.cell(row=r, column=6).value)
            # Audit Phase 2 (re-check) — sejak dosen.nidn punya unique index
            # (partial, mengizinkan banyak NIDN kosong), baris Excel lama yang
            # NIDN-nya kebetulan sama dgn dosen lain (typo entri data, atau
            # dua ejaan nama berbeda utk 1 NIDN yang sama) akan membuat
            # INSERT ini gagal dgn IntegrityError. Beda dari import_generic.py
            # (proses_upload() di sana SUDAH membungkus tiap baris dgn
            # try/except sendiri), loop run_import() ini TIDAK — tanpa
            # try/except di sini, satu baris NIDN bentrok akan MENGGAGALKAN
            # SELURUH proses migrasi Excel di tengah jalan, bukan cuma baris
            # itu. Fallback: dosen tetap dibuat (dicocokkan lewat NAMA tetap
            # jalan utk baris-baris berikutnya yang mereferensikannya lewat
            # find_dosen()), hanya NIDN-nya dikosongkan dan operator diberi
            # tahu lewat ringkasan supaya membetulkan manual.
            try:
                cur = conn.execute(
                    "INSERT INTO dosen(nidn,nama,no_hp,email,aktif) VALUES(?,?,?,?,1)",
                    (nidn, nama, no_hp, email),
                )
            except sqlite3.IntegrityError:
                cur = conn.execute(
                    "INSERT INTO dosen(nidn,nama,no_hp,email,aktif) VALUES(?,?,?,?,1)",
                    ("", nama, no_hp, email),
                )
                n_dosen_nidn_bentrok += 1
            dosen_id_by_name[key] = cur.lastrowid
            n_dosen_baru += 1
        conn.commit()
        ringkasan.append(
            f"{n_dosen_baru} dosen baru diimpor ({len(dosen_id_by_name)} total di database)"
        )
        if n_dosen_nidn_bentrok:
            ringkasan.append(
                f"{n_dosen_nidn_bentrok} dosen baru punya NIDN yang sudah dipakai dosen lain -- "
                "disimpan TANPA NIDN (dosennya tetap masuk, dicocokkan lewat nama), mohon "
                "dicek & dibetulkan manual di halaman Data Dosen."
            )

    def find_dosen(name):
        if not name:
            return None
        return dosen_id_by_name.get(norm_key(name))

    # 3) Mahasiswa — dicocokkan ke DATABASE dulu via NIM (UNIQUE)
    mhs_id_by_nim = {}
    for row in conn.execute("SELECT id, nim FROM mahasiswa"):
        mhs_id_by_nim[row["nim"]] = row["id"]
    n_mhs_baru = 0
    if "Data Mahasiswa" in wb.sheetnames:
        ws = wb["Data Mahasiswa"]
        for r in range(6, ws.max_row + 1):
            nim = norm(ws.cell(row=r, column=2).value)
            nama = norm(ws.cell(row=r, column=3).value)
            if not nim or not nama:
                continue
            if nim in mhs_id_by_nim:
                continue
            jk = norm(ws.cell(row=r, column=4).value)
            tempat_lahir = norm(ws.cell(row=r, column=5).value)
            tgl_lahir = fmt(ws.cell(row=r, column=6).value)
            no_hp = norm(ws.cell(row=r, column=7).value)
            email_nik = norm(ws.cell(row=r, column=8).value)
            status = norm(ws.cell(row=r, column=10).value) or "Aktif"
            status_ta_raw = norm(ws.cell(row=r, column=11).value)
            status_ta = STATUS_TA_MAP.get(status_ta_raw.lower(), C.STATUS_TA_BELUM)
            angkatan = nim[:4] if len(nim) >= 4 and nim[:4].isdigit() else ""
            try:
                cur = conn.execute(
                    "INSERT INTO mahasiswa(nim,nama,jk,tempat_lahir,tgl_lahir,no_hp,email_nik,"
                    "angkatan,status,status_ta,skema) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
                    (
                        nim,
                        nama,
                        jk,
                        tempat_lahir,
                        tgl_lahir,
                        no_hp,
                        email_nik,
                        angkatan,
                        status,
                        status_ta,
                        skema_default,
                    ),
                )
                mhs_id_by_nim[nim] = cur.lastrowid
                n_mhs_baru += 1
            except Exception as e:
                # Audit temuan sistemik — terjemahkan galat mentah ke bahasa
                # manusia lewat error_utils.pesan_ramah_import(). Kalau
                # polanya tidak dikenali (kemungkinan bug, bukan sekadar
                # constraint DB), tetap catat traceback penuh ke log supaya
                # tidak lolos tanpa jejak hanya lewat teks ringkasan impor.
                if EH.pesan_ramah_db(e) is None:
                    EH.logger.exception("Import Excel — gagal proses baris %s (NIM %s)", r, nim)
                ringkasan.append(f"! Lewati baris {r} (NIM {nim}): {EH.pesan_ramah_import(e)}")
        conn.commit()
        ringkasan.append(
            f"{n_mhs_baru} mahasiswa baru diimpor ({len(mhs_id_by_nim)} total di database)"
        )

    # Audit Menyeluruh (temuan verifikasi data produksi nyata, Agustus 2026)
    # — sheet "Data Mahasiswa" institusi ini ternyata TIDAK memuat SEMUA
    # NIM yang muncul di sheet operasional (Pengajuan Judul, Seminar,
    # Sidang, dst.): mahasiswa angkatan lama yang "telat" (baru
    # menyelesaikan skripsi belakangan) sudah tidak tercatat lagi di
    # master "Data Mahasiswa" institusi, padahal riwayat sidang/
    # yudisium/tracer study mereka masih ada. Sebelum perbaikan ini,
    # baris seperti itu DIAM-DIAM DILEWATI (nim not in mhs_id_by_nim ->
    # continue) -- kehilangan seluruh riwayat kelulusan mahasiswa yang
    # sebenarnya sah, tanpa ada tanda apa pun di ringkasan impor.
    #
    # get_or_create_mhs() menggantikan pengecekan itu di 7 sheet turunan
    # di bawah (Pengajuan Judul, Penetapan Pembimbing, Seminar, Sidang,
    # Rencana Yudisium, Wisuda, Tracer Study): kalau NIM belum ada,
    # BUATKAN baris mahasiswa minimal (NIM + nama, field lain kosong,
    # angkatan ditebak dari 4 digit awal NIM) alih-alih membuang
    # riwayatnya. Baris hasil auto-create ini ditandai jelas di kolom
    # `catatan` supaya operator tahu perlu melengkapi data pribadi
    # mahasiswa itu secara manual, dan status_ta-nya dihitung ulang di
    # akhir proses (lihat penutup run_import()) dari riwayat yang
    # sebenarnya terimpor -- bukan ditebak di sini.
    nim_auto_dibuat = {}

    def get_or_create_mhs(nim, nama):
        nim = norm(nim)
        if not nim:
            return None
        if nim in mhs_id_by_nim:
            return mhs_id_by_nim[nim]
        nama = norm(nama) or f"(Nama belum diisi — NIM {nim})"
        angkatan = nim[:4] if len(nim) >= 4 and nim[:4].isdigit() else ""
        cur = conn.execute(
            "INSERT INTO mahasiswa(nim,nama,angkatan,status,status_ta,skema,catatan) "
            "VALUES(?,?,?,?,?,?,?)",
            (
                nim,
                nama,
                angkatan,
                "Aktif",
                C.STATUS_TA_BELUM,
                skema_default,
                "Dibuat otomatis saat impor Excel -- NIM ini muncul di sheet riwayat TA "
                "(Pengajuan/Seminar/Sidang/dll) tapi tidak ada di sheet 'Data Mahasiswa'. "
                "Mohon lengkapi data pribadinya (JK, tempat/tgl lahir, no HP, dll) secara manual.",
            ),
        )
        mhs_id_by_nim[nim] = cur.lastrowid
        nim_auto_dibuat[nim] = nama
        return cur.lastrowid

    # 4) Pengajuan Judul (+ Review Judul digabung berdasarkan Kode Pengajuan)
    review_by_kode = {}
    if "Review Judul" in wb.sheetnames:
        ws = wb["Review Judul"]
        for r in range(7, ws.max_row + 1):
            kode = norm(ws.cell(row=r, column=2).value)
            nim = norm(ws.cell(row=r, column=3).value)
            if not nim and not kode:
                continue
            review_by_kode[kode or nim] = {
                "rev1": norm(ws.cell(row=r, column=9).value),
                "rev2": norm(ws.cell(row=r, column=10).value),
                "rev3": norm(ws.cell(row=r, column=11).value) if ws.max_column >= 11 else "",
                "status_final": (
                    norm(ws.cell(row=r, column=12).value) if ws.max_column >= 12 else ""
                ),
                "tgl_review": fmt(ws.cell(row=r, column=13).value) if ws.max_column >= 13 else "",
                "catatan": norm(ws.cell(row=r, column=14).value) if ws.max_column >= 14 else "",
                "judul_final": norm(ws.cell(row=r, column=15).value) if ws.max_column >= 15 else "",
            }

    n_pengajuan = 0
    if "Pengajuan Judul" in wb.sheetnames:
        ws = wb["Pengajuan Judul"]
        # Audit Menyeluruh (temuan verifikasi data produksi nyata, Agustus 2026)
        # — template "Pengajuan Judul" institusi ini SUDAH BERKEMBANG sejak
        # kode ini pertama ditulis: kolom "Tgl Pengajuan" kini eksplisit di
        # posisi 3 (dulu tidak ada / tergabung), sehingga NIM & seluruh
        # kolom setelahnya bergeser. Sheet "Review Judul" yang berdekatan
        # TIDAK ikut berubah (NIM-nya masih di kolom 3) -- itu sebabnya bug
        # ini HANYA di sheet ini, tidak di sheet lain (sudah diverifikasi
        # sheet lain cocok 1:1 terhadap file produksi nyata). Kolom
        # "tahap" juga TIDAK ADA sebagai kolom per-baris di template ini
        # (beda dari Penetapan Pembimbing/Seminar yang punya kolom
        # "Gelombang"/"Tahap" eksplisit) -- di sini tahap hanya muncul
        # sebagai baris pemisah section "▶ TAHAP 1 – TA. 2025-2026" di atas
        # sekelompok baris. tahap_berjalan melacak baris pemisah terakhir
        # yang terlihat, diterapkan ke semua baris data sesudahnya sampai
        # baris pemisah berikutnya.
        tahap_berjalan = ""
        pola_tahap = re.compile(r"TAHAP\s*(\d+).*?(\d{4})\s*[-–]\s*(\d{4})", re.IGNORECASE)
        for r in range(6, ws.max_row + 1):
            sel_a1 = norm(ws.cell(row=r, column=1).value)
            if sel_a1.startswith("▶"):
                m = pola_tahap.search(sel_a1)
                tahap_berjalan = (
                    f"Tahap {m.group(1)} {m.group(2)} - {m.group(3)}"
                    if m
                    else sel_a1.lstrip("▶").strip()
                )
                continue
            kode = norm(ws.cell(row=r, column=2).value)
            nim = norm(ws.cell(row=r, column=4).value)
            if not nim:
                continue
            mid = get_or_create_mhs(nim, ws.cell(row=r, column=5).value)
            already = conn.execute(
                "SELECT id FROM pengajuan_judul WHERE mahasiswa_id=? AND kode_pengajuan=?",
                (mid, kode),
            ).fetchone()
            if already:
                continue
            rv = review_by_kode.get(kode or nim, {})
            conn.execute(
                "INSERT INTO pengajuan_judul(kode_pengajuan,tgl_pengajuan,mahasiswa_id,semester,"
                "tahap,jml_sks,ipk,judul1,judul2,rev1_ket,rev2_ket,rev3_ket,status_final,"
                "tgl_review,catatan_reviewer,judul_final) "
                "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (
                    kode,
                    fmt(ws.cell(row=r, column=3).value),
                    mid,
                    norm(ws.cell(row=r, column=6).value),
                    tahap_berjalan,
                    norm(ws.cell(row=r, column=8).value),
                    norm(ws.cell(row=r, column=9).value),
                    norm(ws.cell(row=r, column=11).value),
                    norm(ws.cell(row=r, column=12).value),
                    rv.get("rev1", ""),
                    rv.get("rev2", ""),
                    rv.get("rev3", ""),
                    rv.get("status_final") or "Diajukan",
                    rv.get("tgl_review", ""),
                    rv.get("catatan", ""),
                    rv.get("judul_final", ""),
                ),
            )
            n_pengajuan += 1
        conn.commit()
        ringkasan.append(f"{n_pengajuan} pengajuan judul baru diimpor")

    # 5) Penetapan Pembimbing
    n_penetapan = 0
    if "Penetapan Pembimbing" in wb.sheetnames:
        ws = wb["Penetapan Pembimbing"]
        for r in range(7, ws.max_row + 1):
            nim = norm(ws.cell(row=r, column=2).value)
            if not nim:
                continue
            mid = get_or_create_mhs(nim, ws.cell(row=r, column=3).value)
            already = conn.execute(
                "SELECT id FROM penetapan_pembimbing WHERE mahasiswa_id=?", (mid,)
            ).fetchone()
            if already:
                continue
            conn.execute(
                "INSERT INTO penetapan_pembimbing(mahasiswa_id,semester,tahap,judul_final,"
                "pembimbing1_id,pembimbing2_id,tgl_penetapan,no_sk,pembahas1_id,pembahas2_id,"
                "pembahas3_id,ketua_sidang_id,penguji1_id,penguji2_id,penguji3_id,penguji4_id,"
                "link_sk) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (
                    mid,
                    norm(ws.cell(row=r, column=4).value),
                    norm(ws.cell(row=r, column=5).value),
                    norm(ws.cell(row=r, column=7).value),
                    find_dosen(ws.cell(row=r, column=8).value),
                    find_dosen(ws.cell(row=r, column=9).value),
                    fmt(ws.cell(row=r, column=10).value),
                    norm(ws.cell(row=r, column=11).value),
                    find_dosen(ws.cell(row=r, column=12).value),
                    find_dosen(ws.cell(row=r, column=13).value),
                    find_dosen(ws.cell(row=r, column=14).value),
                    find_dosen(ws.cell(row=r, column=15).value),
                    find_dosen(ws.cell(row=r, column=16).value),
                    find_dosen(ws.cell(row=r, column=17).value),
                    find_dosen(ws.cell(row=r, column=18).value),
                    find_dosen(ws.cell(row=r, column=19).value),
                    norm(ws.cell(row=r, column=20).value),
                ),
            )
            n_penetapan += 1
        conn.commit()
        ringkasan.append(f"{n_penetapan} penetapan pembimbing baru diimpor")

    # 6) Seminar
    n_seminar = 0
    n_seminar_status_tidak_dikenali = 0
    if "Seminar" in wb.sheetnames:
        ws = wb["Seminar"]
        for r in range(7, ws.max_row + 1):
            nim = norm(ws.cell(row=r, column=2).value)
            if not nim:
                continue
            mid = get_or_create_mhs(nim, ws.cell(row=r, column=3).value)
            already = conn.execute("SELECT id FROM seminar WHERE mahasiswa_id=?", (mid,)).fetchone()
            if already:
                continue

            def chk(col):
                v = norm(ws.cell(row=r, column=col).value)
                return 1 if v in ("√", "v", "V", "Ya", "YA") else 0

            status_raw = norm(ws.cell(row=r, column=4).value)
            status = _normalisasi_status_seminar(status_raw) if status_raw else "Terdaftar"
            if status_raw and status_raw.lower() not in STATUS_SEMINAR_MAP:
                n_seminar_status_tidak_dikenali += 1

            conn.execute(
                "INSERT INTO seminar(mahasiswa_id,status,tgl_daftar,tgl_seminar,jam,"
                "chk_persetujuan,chk_bukti_bayar,chk_mendeley,chk_krs,chk_bimbingan,"
                "chk_hardcopy,chk_turnitin,judul_diseminarkan,ada_perubahan,penguji_ketua_id,"
                "penguji_anggota1_id,penguji_anggota2_id) "
                "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (
                    mid,
                    status,
                    fmt(ws.cell(row=r, column=8).value),
                    fmt(ws.cell(row=r, column=9).value),
                    fmt(ws.cell(row=r, column=10).value),
                    chk(11),
                    chk(12),
                    chk(13),
                    chk(14),
                    chk(15),
                    chk(16),
                    chk(17),
                    norm(ws.cell(row=r, column=19).value),
                    norm(ws.cell(row=r, column=20).value) or "Tidak",
                    find_dosen(ws.cell(row=r, column=24).value),
                    find_dosen(ws.cell(row=r, column=25).value),
                    find_dosen(ws.cell(row=r, column=26).value),
                ),
            )
            n_seminar += 1
        conn.commit()
        ringkasan.append(f"{n_seminar} data seminar baru diimpor")
        if n_seminar_status_tidak_dikenali:
            ringkasan.append(
                f'{n_seminar_status_tidak_dikenali} baris seminar punya status yang tidak '
                'dikenali (di luar "Terdaftar"/"Selesai"/"Batal") -- disimpan sebagai '
                '"Terdaftar" sementara, mohon dicek & diperbaiki manual di halaman Seminar.'
            )

    # 7) Sidang Skripsi
    n_sidang = 0
    n_sidang_status_tidak_dikenali = 0
    if "Sidang Skripsi" in wb.sheetnames:
        ws = wb["Sidang Skripsi"]
        for r in range(7, ws.max_row + 1):
            nim = norm(ws.cell(row=r, column=2).value)
            if not nim:
                continue
            keterangan_raw = norm(ws.cell(row=r, column=27).value)
            if not keterangan_raw:
                continue
            mid = get_or_create_mhs(nim, ws.cell(row=r, column=3).value)
            # Audit P0 #10 — status_kelulusan HARUS salah satu dari
            # C.STATUS_KELULUSAN_SIDANG (dicocokkan case-insensitive lewat
            # _normalisasi_status_sidang). Kalau teks di Excel tidak
            # dikenali, jangan simpan mentah-mentah (bisa merusak
            # status_sidang_mahasiswa/recalculate_status_ta) — simpan NULL
            # dan minta operator melengkapi manual di halaman Sidang.
            status_kelulusan = _normalisasi_status_sidang(keterangan_raw)
            if status_kelulusan is None:
                n_sidang_status_tidak_dikenali += 1
            existing = conn.execute(
                "SELECT id FROM sidang WHERE mahasiswa_id=? AND keterangan_perubahan=?",
                (mid, norm(ws.cell(row=r, column=12).value)),
            ).fetchone()
            if existing:
                continue
            nilai = ws.cell(row=r, column=25).value
            try:
                nilai = float(nilai) if nilai not in (None, "") else None
            except (TypeError, ValueError):
                nilai = None
            conn.execute(
                "INSERT INTO sidang(mahasiswa_id,tgl_sidang,jam_sidang,judul_sidang,"
                "ada_perubahan,keterangan_perubahan,ketua_id,sekretaris_id,anggota1_id,"
                "anggota2_id,anggota3_id,nilai_angka,status_kelulusan) "
                "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (
                    mid,
                    fmt(ws.cell(row=r, column=6).value),
                    fmt(ws.cell(row=r, column=7).value),
                    norm(ws.cell(row=r, column=10).value),
                    norm(ws.cell(row=r, column=11).value) or "Tidak",
                    norm(ws.cell(row=r, column=12).value),
                    find_dosen(ws.cell(row=r, column=15).value),
                    find_dosen(ws.cell(row=r, column=16).value),
                    find_dosen(ws.cell(row=r, column=17).value),
                    find_dosen(ws.cell(row=r, column=18).value),
                    find_dosen(ws.cell(row=r, column=19).value),
                    nilai,
                    status_kelulusan,
                ),
            )
            n_sidang += 1
        conn.commit()
        ringkasan.append(f"{n_sidang} data sidang baru diimpor")
        if n_sidang_status_tidak_dikenali:
            ringkasan.append(
                f'{n_sidang_status_tidak_dikenali} baris sidang punya status kelulusan yang '
                'tidak dikenali (di luar "LULUS"/"TIDAK LULUS"/"TUNDA") -- kolom Status '
                "Kelulusan disimpan KOSONG, mohon dilengkapi manual di halaman Sidang."
            )

    # 8) & 9) Yudisium / Wisuda: baris auto-generate lewat logic.py
    import app.logic as L

    L.sync_yudisium_dari_sidang(conn)
    n_yud = 0
    if "Rencana Yudisium" in wb.sheetnames:
        ws = wb["Rencana Yudisium"]
        for r in range(7, ws.max_row + 1):
            nim = norm(ws.cell(row=r, column=2).value)
            if not nim:
                continue
            mid = get_or_create_mhs(nim, ws.cell(row=r, column=3).value)
            ipk = ws.cell(row=r, column=8).value
            try:
                ipk = float(ipk) if ipk not in (None, "") else None
            except (TypeError, ValueError):
                ipk = None
            tgl_yud = fmt(ws.cell(row=r, column=11).value)
            no_sk = norm(ws.cell(row=r, column=12).value)
            status_yud = norm(ws.cell(row=r, column=13).value) or "Direncanakan"
            conn.execute(
                "UPDATE yudisium SET ipk_final=?, tgl_yudisium=?, no_sk=?, status_yudisium=? "
                "WHERE mahasiswa_id=?",
                (ipk, tgl_yud, no_sk, status_yud, mid),
            )
            n_yud += 1
        conn.commit()
        ringkasan.append(f"{n_yud} data yudisium dilengkapi (IPK/tgl/SK)")

    L.sync_wisuda_dari_yudisium(conn)
    n_wis = 0
    if "Wisuda" in wb.sheetnames:
        ws = wb["Wisuda"]
        for r in range(7, ws.max_row + 1):
            nim = norm(ws.cell(row=r, column=2).value)
            if not nim:
                continue
            mid = get_or_create_mhs(nim, ws.cell(row=r, column=3).value)
            tgl_w = fmt(ws.cell(row=r, column=11).value)
            no_ij = norm(ws.cell(row=r, column=12).value)
            catatan = norm(ws.cell(row=r, column=13).value)
            conn.execute(
                "UPDATE wisuda SET tgl_wisuda=?, no_ijazah=?, catatan=? WHERE mahasiswa_id=?",
                (tgl_w, no_ij, catatan, mid),
            )
            n_wis += 1
        conn.commit()
        ringkasan.append(f"{n_wis} data wisuda dilengkapi")

    # 10) Tracer Study Alumni
    n_tracer = 0
    if "Tracer Study Alumni" in wb.sheetnames:
        ws = wb["Tracer Study Alumni"]
        for r in range(8, ws.max_row + 1):
            nim = norm(ws.cell(row=r, column=2).value)
            if not nim:
                continue
            mid = get_or_create_mhs(nim, ws.cell(row=r, column=3).value)
            already = conn.execute(
                "SELECT id FROM tracer_study WHERE mahasiswa_id=?", (mid,)
            ).fetchone()
            if already:
                continue
            conn.execute(
                "INSERT INTO tracer_study(mahasiswa_id,status_saat_ini,nama_instansi,posisi,"
                "kesesuaian_bidang,waktu_tunggu,studi_lanjut,program_lanjut,no_hp,catatan) "
                "VALUES(?,?,?,?,?,?,?,?,?,?)",
                (
                    mid,
                    norm(ws.cell(row=r, column=9).value),
                    norm(ws.cell(row=r, column=10).value),
                    norm(ws.cell(row=r, column=11).value),
                    norm(ws.cell(row=r, column=12).value),
                    norm(ws.cell(row=r, column=13).value),
                    norm(ws.cell(row=r, column=14).value),
                    norm(ws.cell(row=r, column=15).value),
                    norm(ws.cell(row=r, column=16).value),
                    norm(ws.cell(row=r, column=17).value),
                ),
            )
            n_tracer += 1
        conn.commit()
        ringkasan.append(f"{n_tracer} data tracer study baru diimpor")

    # Audit Menyeluruh (temuan verifikasi data produksi nyata) — tutup
    # proses get_or_create_mhs(): status_ta mahasiswa yang auto-dibuat
    # dihitung ulang di SINI (setelah SEMUA sheet riwayat TA selesai
    # diimpor), bukan ditebak saat baris mahasiswa-nya dibuat -- supaya
    # mencerminkan riwayat SEBENARNYA yang baru saja masuk (mis. mahasiswa
    # yang ternyata punya sidang LULUS harus berakhir status_ta=LULUS,
    # bukan tertinggal di default "Belum Mengajukan Judul").
    if nim_auto_dibuat:
        for nim in nim_auto_dibuat:
            L.recalculate_status_ta(
                conn, mhs_id_by_nim[nim], dipicu_oleh="Impor Excel (mahasiswa auto-dibuat)"
            )
        conn.commit()
        daftar = ", ".join(f"{nama} ({nim})" for nim, nama in list(nim_auto_dibuat.items())[:10])
        lebih = f", dan {len(nim_auto_dibuat) - 10} lainnya" if len(nim_auto_dibuat) > 10 else ""
        ringkasan.append(
            f"⚠ {len(nim_auto_dibuat)} mahasiswa OTOMATIS DIBUAT karena NIM-nya muncul di riwayat "
            f"TA tapi tidak ada di sheet 'Data Mahasiswa': {daftar}{lebih}. Data pribadinya "
            "(JK, tempat/tgl lahir, dll.) masih kosong -- mohon dilengkapi manual lewat halaman "
            "Data Mahasiswa kalau diperlukan."
        )

    return ringkasan


# =============================================================================
# Audit Menyeluruh — Migrasi data RPL (Rekognisi Pembelajaran Lampau)
# =============================================================================
# Workbook "Aplikasi Manajemen RPL" institusi ini strukturnya BERBEDA TOTAL
# dari workbook skripsi reguler (run_import() di atas) -- bukan variasi
# kecil, jadi sengaja dibuat fungsi TERPISAH, bukan dipaksa masuk ke
# run_import() yang sudah ada. Perbedaan paling mendasar: "Pengajuan Judul"
# dan "Penetapan Pembimbing" digabung jadi SATU sheet ("JUDUL & PEMBIMBING")
# di sini, bukan dua sheet terpisah.
#
# Mahasiswa yang masuk lewat importer ini SELALU diberi skema='RPL' (bukan
# opsional seperti skema_default di run_import()) -- itulah esensi
# migrasi ini, seluruh isi file memang representasi jalur RPL.
def run_import_rpl(xlsx_path, conn):
    """Migrasi dari workbook 'Aplikasi Manajemen RPL' ke database SQLite.
    Mengembalikan list baris ringkasan (string), sama seperti run_import()."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ringkasan = []

    # 1) Dosen — pola SAMA PERSIS dengan sheet "Data Dosen" di workbook
    # skripsi reguler (header baris 6, data baris 7+, kolom NIDN/Nama/
    # ProgramStudi/No.HP/Email) -- institusi ini memakai template dosen yang
    # konsisten di kedua workbook.
    dosen_id_by_name = {}
    for row in conn.execute("SELECT id, nama FROM dosen"):
        dosen_id_by_name[norm_key(row["nama"])] = row["id"]
    n_dosen_baru = 0
    if "DOSEN" in wb.sheetnames:
        ws = wb["DOSEN"]
        for r in range(7, ws.max_row + 1):
            nama = norm(ws.cell(row=r, column=3).value)
            if not nama:
                continue
            key = norm_key(nama)
            if key in dosen_id_by_name:
                continue
            nidn = norm(ws.cell(row=r, column=2).value)
            no_hp = norm(ws.cell(row=r, column=5).value)
            email = norm(ws.cell(row=r, column=6).value)
            try:
                cur = conn.execute(
                    "INSERT INTO dosen(nidn,nama,no_hp,email,aktif,status_kepegawaian) "
                    "VALUES(?,?,?,?,1,'Aktif')",
                    (nidn, nama, no_hp, email),
                )
            except sqlite3.IntegrityError:
                # NIDN kembar dgn dosen lain (mis. dosen yg sama sudah
                # terimpor dari workbook skripsi reguler dgn ejaan nama
                # sedikit beda) -- simpan tanpa NIDN drpd gagal total,
                # sama seperti fallback di blok Data Dosen run_import().
                cur = conn.execute(
                    "INSERT INTO dosen(nidn,nama,no_hp,email,aktif,status_kepegawaian) "
                    "VALUES('',?,?,?,1,'Aktif')",
                    (nama, no_hp, email),
                )
            dosen_id_by_name[key] = cur.lastrowid
            n_dosen_baru += 1
        conn.commit()
        ringkasan.append(f"{n_dosen_baru} dosen baru diimpor ({len(dosen_id_by_name)} total di database)")

    def find_dosen(nama):
        nama = norm(nama)
        if not nama:
            return None
        row = dosen_id_by_name.get(norm_key(nama))
        return row

    # 2) Data Induk -> mahasiswa (skema='RPL', SELALU, bukan opsional)
    mhs_id_by_nim = {}
    for row in conn.execute("SELECT id, nim FROM mahasiswa"):
        mhs_id_by_nim[row["nim"]] = row["id"]
    n_mhs_baru = 0
    if "DATA INDUK" in wb.sheetnames:
        ws = wb["DATA INDUK"]
        for r in range(4, ws.max_row + 1):
            npm = norm(ws.cell(row=r, column=2).value)
            nama = norm(ws.cell(row=r, column=3).value)
            if not npm or not nama or npm in mhs_id_by_nim:
                continue
            jk_raw = norm(ws.cell(row=r, column=7).value)
            jk = "L" if jk_raw.lower().startswith("laki") else ("P" if jk_raw else "")
            tempat_lahir = norm(ws.cell(row=r, column=8).value)
            no_hp = norm(ws.cell(row=r, column=9).value)
            status = norm(ws.cell(row=r, column=14).value) or "Aktif"
            if status not in C.STATUS_MHS_LIST:
                status = "Aktif"
            jalur = norm(ws.cell(row=r, column=5).value)
            angkatan = npm[:4] if len(npm) >= 4 and npm[:4].isdigit() else ""
            cur = conn.execute(
                "INSERT INTO mahasiswa(nim,nama,jk,tempat_lahir,no_hp,angkatan,status,"
                "status_ta,skema,catatan) VALUES(?,?,?,?,?,?,?,?,?,?)",
                (
                    npm, nama, jk, tempat_lahir, no_hp, angkatan, status,
                    C.STATUS_TA_BELUM, "RPL",
                    f"Jalur: {jalur}" if jalur else "",
                ),
            )
            mhs_id_by_nim[npm] = cur.lastrowid
            n_mhs_baru += 1
        conn.commit()
        ringkasan.append(f"{n_mhs_baru} mahasiswa RPL baru diimpor ({len(mhs_id_by_nim)} total di database)")

    def get_or_create_mhs(npm, nama):
        npm = norm(npm)
        if not npm:
            return None
        if npm in mhs_id_by_nim:
            return mhs_id_by_nim[npm]
        nama = norm(nama) or f"(Nama belum diisi — NPM {npm})"
        angkatan = npm[:4] if len(npm) >= 4 and npm[:4].isdigit() else ""
        cur = conn.execute(
            "INSERT INTO mahasiswa(nim,nama,angkatan,status,status_ta,skema,catatan) "
            "VALUES(?,?,?,?,?,?,?)",
            (
                npm, nama, angkatan, "Aktif", C.STATUS_TA_BELUM, "RPL",
                "Dibuat otomatis saat impor RPL -- NPM ini muncul di sheet riwayat TA tapi "
                "tidak ada di sheet 'DATA INDUK'. Mohon lengkapi data pribadinya secara manual.",
            ),
        )
        mhs_id_by_nim[npm] = cur.lastrowid
        return cur.lastrowid

    # 3) Judul & Pembimbing -> pengajuan_judul + penetapan_pembimbing
    # (SATU sheet di workbook RPL ini, beda dari workbook reguler yang
    # memisahkannya jadi 2 sheet -- lihat komentar modul di atas).
    n_pengajuan = n_penetapan = 0
    if "JUDUL & PEMBIMBING" in wb.sheetnames:
        ws = wb["JUDUL & PEMBIMBING"]
        for r in range(5, ws.max_row + 1):
            npm = norm(ws.cell(row=r, column=2).value)
            if not npm:
                continue
            mid = get_or_create_mhs(npm, ws.cell(row=r, column=3).value)
            judul = norm(ws.cell(row=r, column=4).value)
            status_judul = norm(ws.cell(row=r, column=8).value)
            tahap_seminar = norm(ws.cell(row=r, column=16).value)

            kode = f"RPL-{npm}"
            already = conn.execute(
                "SELECT id FROM pengajuan_judul WHERE mahasiswa_id=? AND kode_pengajuan=?",
                (mid, kode),
            ).fetchone()
            if not already and judul:
                conn.execute(
                    "INSERT INTO pengajuan_judul(kode_pengajuan,tgl_pengajuan,mahasiswa_id,"
                    "tahap,judul1,status_final,judul_final) VALUES(?,?,?,?,?,?,?)",
                    (
                        kode,
                        fmt(ws.cell(row=r, column=6).value),
                        mid,
                        tahap_seminar,
                        judul,
                        "Disetujui" if status_judul == "Disetujui" else (status_judul or "Diajukan"),
                        judul if status_judul == "Disetujui" else "",
                    ),
                )
                n_pengajuan += 1

            pembimbing1 = norm(ws.cell(row=r, column=10).value)
            already_pp = conn.execute(
                "SELECT id FROM penetapan_pembimbing WHERE mahasiswa_id=?", (mid,)
            ).fetchone()
            if not already_pp and pembimbing1:
                conn.execute(
                    "INSERT INTO penetapan_pembimbing(mahasiswa_id,tahap,judul_final,"
                    "pembimbing1_id,pembimbing2_id,penguji1_id,penguji2_id) "
                    "VALUES(?,?,?,?,?,?,?)",
                    (
                        mid,
                        tahap_seminar,
                        judul,
                        find_dosen(pembimbing1),
                        find_dosen(ws.cell(row=r, column=11).value),
                        find_dosen(ws.cell(row=r, column=12).value),
                        find_dosen(ws.cell(row=r, column=13).value),
                    ),
                )
                n_penetapan += 1
        conn.commit()
        ringkasan.append(f"{n_pengajuan} pengajuan judul RPL baru diimpor")
        ringkasan.append(f"{n_penetapan} penetapan pembimbing RPL baru diimpor")

    # 4) Seminar Proposal -> seminar
    n_seminar = 0
    if "SEMINAR PROPOSAL" in wb.sheetnames:
        ws = wb["SEMINAR PROPOSAL"]
        for r in range(4, ws.max_row + 1):
            npm = norm(ws.cell(row=r, column=2).value)
            if not npm:
                continue
            already = conn.execute(
                "SELECT id FROM seminar WHERE mahasiswa_id=(SELECT id FROM mahasiswa WHERE nim=?)",
                (npm,),
            ).fetchone()
            if already:
                continue
            mid = get_or_create_mhs(npm, ws.cell(row=r, column=3).value)
            status = _normalisasi_status_seminar(ws.cell(row=r, column=4).value)
            conn.execute(
                "INSERT INTO seminar(mahasiswa_id,status,tgl_seminar,jam,judul_diseminarkan,"
                "tahap,penguji_ketua_id,penguji_anggota1_id,penguji_anggota2_id) "
                "VALUES(?,?,?,?,?,?,?,?,?)",
                (
                    mid,
                    status,
                    fmt(ws.cell(row=r, column=10).value),
                    norm(ws.cell(row=r, column=11).value),
                    norm(ws.cell(row=r, column=6).value),
                    norm(ws.cell(row=r, column=9).value),
                    find_dosen(ws.cell(row=r, column=12).value),
                    find_dosen(ws.cell(row=r, column=13).value),
                    find_dosen(ws.cell(row=r, column=14).value),
                ),
            )
            n_seminar += 1
        conn.commit()
        ringkasan.append(f"{n_seminar} data seminar RPL baru diimpor")

    # 5) Sidang Skripsi -> sidang (sheet mungkin masih kosong utk kohort
    # yang belum sampai tahap sidang -- importer tetap disiapkan supaya
    # langsung bisa dipakai begitu datanya terisi, tanpa perlu update lagi).
    n_sidang = 0
    if "SIDANG SKRIPSI" in wb.sheetnames:
        ws = wb["SIDANG SKRIPSI"]
        for r in range(4, ws.max_row + 1):
            npm = norm(ws.cell(row=r, column=2).value)
            if not npm:
                continue
            status_kelulusan = _normalisasi_status_sidang(ws.cell(row=r, column=14).value)
            if status_kelulusan is None:
                continue
            mid = get_or_create_mhs(npm, ws.cell(row=r, column=3).value)
            conn.execute(
                "INSERT INTO sidang(mahasiswa_id,judul_sidang,tgl_sidang,ketua_id,sekretaris_id,"
                "anggota1_id,anggota2_id,anggota3_id,nilai_angka,status_kelulusan) "
                "VALUES(?,?,?,?,?,?,?,?,?,?)",
                (
                    mid,
                    norm(ws.cell(row=r, column=4).value),
                    fmt(ws.cell(row=r, column=8).value),
                    find_dosen(ws.cell(row=r, column=9).value),
                    find_dosen(ws.cell(row=r, column=10).value),
                    find_dosen(ws.cell(row=r, column=11).value),
                    find_dosen(ws.cell(row=r, column=12).value),
                    find_dosen(ws.cell(row=r, column=13).value),
                    ws.cell(row=r, column=15).value,
                    status_kelulusan,
                ),
            )
            n_sidang += 1
        conn.commit()
        ringkasan.append(f"{n_sidang} data sidang RPL baru diimpor")

    # 6) Yudisium & Wisuda -> yudisium (UPDATE, baris dibuat otomatis dari
    # sidang LULUS lewat logic.sync_yudisium_dari_sidang di run_import())
    # + wisuda. Sheet gabungan di sini, beda dari workbook reguler yang
    # memisahkannya jadi 2 sheet.
    import app.logic as L

    L.sync_yudisium_dari_sidang(conn)
    n_yud = n_wis = 0
    if "YUDISIUM & WISUDA" in wb.sheetnames:
        ws = wb["YUDISIUM & WISUDA"]
        for r in range(4, ws.max_row + 1):
            npm = norm(ws.cell(row=r, column=2).value)
            if not npm or npm not in mhs_id_by_nim:
                continue
            mid = mhs_id_by_nim[npm]
            ipk = ws.cell(row=r, column=6).value
            status_yud = norm(ws.cell(row=r, column=8).value)
            if status_yud not in C.STATUS_YUDISIUM_LIST:
                status_yud = None
            cur = conn.execute(
                "UPDATE yudisium SET ipk_final=COALESCE(?,ipk_final), "
                "tgl_yudisium=COALESCE(NULLIF(?,''),tgl_yudisium), "
                "status_yudisium=COALESCE(?,status_yudisium) WHERE mahasiswa_id=?",
                (ipk, fmt(ws.cell(row=r, column=10).value), status_yud, mid),
            )
            if cur.rowcount:
                n_yud += 1
            no_ijazah = norm(ws.cell(row=r, column=12).value)
            tgl_wisuda = fmt(ws.cell(row=r, column=11).value)
            if no_ijazah or tgl_wisuda:
                cur2 = conn.execute(
                    "UPDATE wisuda SET tgl_wisuda=COALESCE(NULLIF(?,''),tgl_wisuda), "
                    "no_ijazah=COALESCE(NULLIF(?,''),no_ijazah) WHERE mahasiswa_id=?",
                    (tgl_wisuda, no_ijazah, mid),
                )
                if cur2.rowcount:
                    n_wis += 1
        conn.commit()
        L.sync_wisuda_dari_yudisium(conn)
        conn.commit()
        ringkasan.append(f"{n_yud} data yudisium RPL dilengkapi (IPK/tgl/status)")
        ringkasan.append(f"{n_wis} data wisuda RPL dilengkapi")

    # Audit Menyeluruh — beda dari run_import() (yang mengisi status_ta
    # awal dari kolom "Status Pengajuan TA" eksplisit di sheet Data
    # Mahasiswa), DATA INDUK RPL TIDAK punya kolom status_ta -- semua
    # mahasiswa RPL di atas mulai dari default STATUS_TA_BELUM. Hitung
    # ulang di sini utk SELURUH mahasiswa RPL yang tersentuh (bukan cuma
    # yang auto-dibuat) supaya status_ta-nya mencerminkan riwayat
    # pengajuan/pembimbing/seminar/sidang yang BENAR-BENAR baru masuk.
    for npm, mid in mhs_id_by_nim.items():
        L.recalculate_status_ta(conn, mid, dipicu_oleh="Impor Excel RPL")
    conn.commit()

    return ringkasan


def main():
    """Wrapper CLI (kompatibel dengan pemakaian versi desktop)."""
    if len(sys.argv) < 2:
        print("Pemakaian: python import_excel.py <file.xlsx> [path_db_tujuan]")
        sys.exit(1)
    xlsx_path = sys.argv[1]
    db_path = sys.argv[2] if len(sys.argv) > 2 else db.get_default_db_path()
    conn = db.connect(db_path)
    print(f"→ Basis data tujuan: {db_path}")
    for line in run_import(xlsx_path, conn):
        print("✓ " + line)
    print("\nSelesai.")
    conn.close()


if __name__ == "__main__":
    main()
