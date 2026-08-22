# -*- coding: utf-8 -*-
"""routes/rekap.py — Rekap Pembimbing, Rekap Status, RKP Seminar, RKP
Sidang, Rasio Dosen, Statistik. Semua read-only, memakai fungsi logic.py
yang sama persis dengan versi desktop (logikanya tidak diubah)."""

import io

from flask import Blueprint, current_app, render_template, request, send_file

from app import db as _db
from app import export_utils
from app import logic as L

bp = Blueprint("rekap", __name__, url_prefix="/rekap")


def _tahap_opsi(conn):
    """Daftar opsi filter tahap untuk rekap — dinamis dari tahap_pengajuan
    (Audit poin 2, klarifikasi a), jatuh balik ke nama_tahap_1/2 lama kalau
    prodi belum pernah membuka tahun ajaran lewat wizard baru."""
    tahap_rows = list(_db.get_tahap_list(conn))
    if tahap_rows:
        return ["Semua"] + [t["nama"] for t in tahap_rows]
    lama = [_db.get_setting(conn, "nama_tahap_1"), _db.get_setting(conn, "nama_tahap_2")]
    return ["Semua"] + [t for t in lama if t]


# Audit Lanjutan (Kelulusan/Tracer Study) — _kirim_excel() sebelumnya
# digandakan persis sama di file ini & routes/kelulusan.py. Sekarang
# dipusatkan ke app/export_utils.py (dipakai lewat alias di bawah supaya
# semua pemanggilan _kirim_excel(...) yang sudah ada di file ini tidak
# perlu diubah satu-satu).
_kirim_excel = export_utils.kirim_excel


# ------------------------------------------------------------- Rekap Pembimbing
def _rows_pembimbing(conn, tahap):
    data = L.rekap_pembimbing(conn, tahap if tahap and tahap != "Semua" else None)
    out = []
    for dosen_id, v in sorted(data.items(), key=lambda kv: kv[1]["nama"]):
        for peran, label in (("pembimbing_1", "Pembimbing 1"), ("pembimbing_2", "Pembimbing 2")):
            for m in v[peran]:
                out.append(
                    (v["nama"], label, m["nim"], m["nama"], m["status_seminar"], m["status_sidang"])
                )
    return out


@bp.route("/pembimbing")
def pembimbing():
    conn = current_app.get_db()
    tahap = request.args.get("tahap", "Semua")
    rows = _rows_pembimbing(conn, tahap)
    return render_template(
        "rekap_pembimbing.html", rows=rows, tahap=tahap, tahap_opsi=_tahap_opsi(conn)
    )


@bp.route("/pembimbing/ekspor")
def pembimbing_ekspor():
    conn = current_app.get_db()
    tahap = request.args.get("tahap", "Semua")
    rows = _rows_pembimbing(conn, tahap)
    return _kirim_excel(
        "Rekap Pembimbing",
        ["Dosen", "Peran", "NIM", "Nama Mahasiswa", "Status Seminar", "Status Sidang"],
        rows,
    )


# ----------------------------------------------------------------- Rekap Status
@bp.route("/status")
def status():
    conn = current_app.get_db()
    d = L.rekap_status_mahasiswa(conn)
    return render_template("rekap_status.html", d=d)


@bp.route("/status/ekspor")
def status_ekspor():
    conn = current_app.get_db()
    d = L.rekap_status_mahasiswa(conn)
    rows = [(x["nim"], x["nama"], x["status_seminar"], x["status_sidang"]) for x in d["detail"]]
    return _kirim_excel(
        "Rekap Status Mahasiswa", ["NIM", "Nama", "Status Seminar", "Status Sidang"], rows
    )


# ------------------------------------------------------------------ RKP Seminar
@bp.route("/rkp-seminar")
def rkp_seminar():
    conn = current_app.get_db()
    tahap = request.args.get("tahap", "Semua")
    rows = L.rkp_seminar(conn, tahap if tahap != "Semua" else None)
    total = sum(r["honor"] for r in rows)
    return render_template(
        "rkp_seminar.html", rows=rows, total=total, tahap=tahap, tahap_opsi=_tahap_opsi(conn)
    )


@bp.route("/rkp-seminar/ekspor")
def rkp_seminar_ekspor():
    conn = current_app.get_db()
    tahap = request.args.get("tahap", "Semua")
    rows = L.rkp_seminar(conn, tahap if tahap != "Semua" else None)
    data = [(r["nama"], r["jumlah"], r["honor"]) for r in rows]
    return _kirim_excel("RKP Seminar", ["Nama Dosen", "Jumlah Peran", "Honor (Rp)"], data)


# ------------------------------------------------------------------- RKP Sidang
@bp.route("/rkp-sidang")
def rkp_sidang():
    conn = current_app.get_db()
    tahap = request.args.get("tahap", "Semua")
    d = L.rkp_sidang(conn, tahap if tahap != "Semua" else None)
    return render_template("rkp_sidang.html", d=d, tahap=tahap, tahap_opsi=_tahap_opsi(conn))


@bp.route("/rkp-sidang/ekspor")
def rkp_sidang_ekspor():
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    conn = current_app.get_db()
    tahap = request.args.get("tahap", "Semua")
    d = L.rkp_sidang(conn, tahap if tahap != "Semua" else None)
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Bagian 1 - Penguji"
    ws1.append(["Nama Dosen", "Jumlah Peran", "Honor (Rp)"])
    for r in d["bagian1"]:
        ws1.append([r["nama"], r["jumlah"], r["honor"]])
    ws2 = wb.create_sheet("Bagian 2 - Pembimbing")
    ws2.append(["Nama Dosen", "Sbg Pembimbing 1", "Sbg Pembimbing 2", "Honor (Rp)"])
    for r in d["bagian2"]:
        ws2.append([r["nama"], r["sbg_pemb1"], r["sbg_pemb2"], r["honor"]])
    for ws in (ws1, ws2):
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1E3A5F")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="RKP_Sidang.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ----------------------------------------------------- Rekap Honor Keuangan
# Audit Modul Pelaksanaan — RKP Seminar & RKP Sidang di atas masing-masing
# menyajikan honornya sendiri-sendiri (dan Bagian 1/2 terpisah lagi untuk
# Sidang), sehingga bagian keuangan harus menjumlahkan sendiri dari 2-3
# laporan berbeda untuk tahu total honor 1 dosen pada 1 tahap. Halaman ini
# menggabungkan SEMUA kategori honor (Seminar-Penguji, Sidang-Penguji,
# Sidang-Pembimbing 1, Sidang-Pembimbing 2) jadi satu tabel per
# dosen x tahap x kategori, siap dipakai sebagai lampiran pencairan honor.
@bp.route("/honor-keuangan")
def honor_keuangan():
    conn = current_app.get_db()
    tahap = request.args.get("tahap", "Semua")
    rows, total = L.rekap_honor_keuangan(conn, tahap if tahap != "Semua" else None)
    per_dosen = {}
    for r in rows:
        per_dosen.setdefault(r["nama"], {"nama": r["nama"], "rows": [], "subtotal": 0.0})
        per_dosen[r["nama"]]["rows"].append(r)
        per_dosen[r["nama"]]["subtotal"] += r["honor"]
    kelompok = [per_dosen[k] for k in sorted(per_dosen.keys())]
    return render_template(
        "rekap_honor_keuangan.html",
        kelompok=kelompok,
        total=total,
        tahap=tahap,
        tahap_opsi=_tahap_opsi(conn),
    )


@bp.route("/honor-keuangan/ekspor")
def honor_keuangan_ekspor():
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    conn = current_app.get_db()
    tahap = request.args.get("tahap", "Semua")
    rows, total = L.rekap_honor_keuangan(conn, tahap if tahap != "Semua" else None)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rekap Honor Keuangan"
    ws.append(["Nama Dosen", "Tahap", "Kategori", "Jumlah", "Tarif (Rp)", "Honor (Rp)"])
    for r in rows:
        ws.append([r["nama"], r["tahap"], r["kategori"], r["jumlah"], r["tarif"], r["honor"]])
    ws.append([])
    ws.append(["", "", "", "", "TOTAL", total])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1E3A5F")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="Rekap_Honor_Keuangan.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ----------------------------------------------------------------- Rasio Dosen
def _rows_rasio(conn, hanya_homebase=True):
    data = L.rekap_rasio_dosen(conn, hanya_homebase=hanya_homebase)
    return data


@bp.route("/rasio-dosen")
def rasio_dosen():
    conn = current_app.get_db()
    # Audit poin 3: default hanya dosen homebase (basis rasio resmi
    # BAN-PT/LAM & PDDikti/BKD) — ?semua=1 menampilkan gabungan dgn dosen luar.
    hanya_homebase = request.args.get("semua") != "1"
    rows = _rows_rasio(conn, hanya_homebase=hanya_homebase)
    return render_template("rasio_dosen.html", rows=rows, hanya_homebase=hanya_homebase)


@bp.route("/rasio-dosen/ekspor")
def rasio_dosen_ekspor():
    conn = current_app.get_db()
    rows = _rows_rasio(conn)
    data = [
        (
            r["nama"],
            r["pemb1"],
            r["pemb2"],
            r["total_bimb"],
            r["sudah_seminar"],
            r["sudah_sidang"],
            r["total_tugas_penguji"],
            r["total_keterlibatan"],
            r["status"],
        )
        for r in rows
    ]
    return _kirim_excel(
        "Rekap Rasio Dosen",
        [
            "Nama Dosen",
            "Pem.1",
            "Pem.2",
            "Total Bimb.",
            "Sudah Seminar",
            "Sudah Sidang",
            "Tugas Penguji",
            "Total Keterlibatan",
            "Status",
        ],
        data,
    )


# --------------------------------------------------------- Rekap Kinerja Dosen (SDM)
# Bab SDM di Rekap & Laporan — disebut di README sbg "menyusul" sejak Modul 4
# (SDM & Kinerja Dosen) dibangun. Dashboard di sdm.index() sudah menghitung
# kesiapan BKD/SISTER per dosen (lihat routes/sdm.py:_hitung_kesiapan), bab
# ini memakai rumus & pola perhitungan yang SAMA PERSIS (tidak
# didup­likasi/ditulis ulang) supaya angka antara Dashboard SDM dan Rekap
# selalu konsisten, lalu menambahkan yang belum ada di sana: rincian jumlah
# entri per 7 kategori log + kemampuan ekspor Excel.
_SDM_KATEGORI_TABEL = [
    ("aktivitas_pendidikan", "Pendidikan & Pengajaran"),
    ("aktivitas_penelitian", "Penelitian"),
    ("aktivitas_pkm", "PKM"),
    ("aktivitas_penunjang", "Penunjang"),
    ("luaran_dosen", "Luaran"),
    ("peran_akademik_dosen", "Peran Akademik"),
    ("timeline_karier_dosen", "Timeline Karier"),
]
# timeline_karier_dosen tidak punya kolom tahun_akademik (peristiwa karier
# seperti kenaikan pangkat/SK tidak terikat 1 periode akademik tertentu,
# lihat db.py) — dikecualikan dari filter & daftar dropdown Tahun Akademik,
# tapi tetap ikut dihitung jumlahnya per dosen (filter tahun tidak berlaku
# utknya, selalu tampil apa adanya).
_SDM_TABEL_BERTAHUN = [t for t, _l in _SDM_KATEGORI_TABEL if t != "timeline_karier_dosen"]


def _rows_kinerja_dosen(conn, hanya_homebase=True, tahun_akademik=None):
    from app.routes.sdm import _hitung_kesiapan, _hitung_reminder

    q = "SELECT * FROM dosen WHERE aktif=1"
    if hanya_homebase:
        q += " AND (status_homebase IS NULL OR status_homebase='Homebase')"
    q += " ORDER BY nama"
    dosen_list = conn.execute(q).fetchall()
    out = []
    for d in dosen_list:
        kesiapan_bkd, kesiapan_sister = _hitung_kesiapan(conn, d["id"])
        jumlah = {}
        for table, _label in _SDM_KATEGORI_TABEL:
            sql = f"SELECT COUNT(*) FROM {table} WHERE dosen_id=?"
            params = [d["id"]]
            if tahun_akademik and table in _SDM_TABEL_BERTAHUN:
                sql += " AND tahun_akademik=?"
                params.append(tahun_akademik)
            jumlah[table] = conn.execute(sql, params).fetchone()[0]
        n_reminder = len(_hitung_reminder(conn, d["id"]))
        out.append(
            {
                "dosen": d,
                "nama": d["nama"],
                "status_homebase": d["status_homebase"] or "Homebase",
                "kesiapan_bkd": kesiapan_bkd,
                "kesiapan_sister": kesiapan_sister,
                "jumlah": jumlah,
                "total_entri": sum(jumlah.values()),
                "n_reminder": n_reminder,
            }
        )
    return out


@bp.route("/kinerja-dosen")
def kinerja_dosen():
    conn = current_app.get_db()
    hanya_homebase = request.args.get("semua") != "1"
    tahun_akademik = request.args.get("tahun_akademik", "") or None
    rows = _rows_kinerja_dosen(conn, hanya_homebase=hanya_homebase, tahun_akademik=tahun_akademik)
    tahun_list = set()
    for table in _SDM_TABEL_BERTAHUN:
        for r in conn.execute(
            f"SELECT DISTINCT tahun_akademik FROM {table} WHERE tahun_akademik IS NOT NULL AND tahun_akademik != ''"
        ).fetchall():
            tahun_list.add(r["tahun_akademik"])
    tahun_list = sorted(tahun_list, reverse=True)
    return render_template(
        "rekap_kinerja_dosen.html",
        rows=rows,
        hanya_homebase=hanya_homebase,
        kategori=_SDM_KATEGORI_TABEL,
        tahun_akademik=tahun_akademik or "",
        tahun_list=tahun_list,
    )


@bp.route("/kinerja-dosen/ekspor")
def kinerja_dosen_ekspor():
    conn = current_app.get_db()
    hanya_homebase = request.args.get("semua") != "1"
    tahun_akademik = request.args.get("tahun_akademik", "") or None
    rows = _rows_kinerja_dosen(conn, hanya_homebase=hanya_homebase, tahun_akademik=tahun_akademik)
    headers = (
        ["Nama Dosen", "Homebase"]
        + [lbl for _t, lbl in _SDM_KATEGORI_TABEL]
        + ["Total Entri", "Kesiapan BKD (%)", "Kesiapan SISTER (%)", "Reminder Aktif"]
    )
    data = []
    for r in rows:
        data.append(
            (
                r["nama"],
                r["status_homebase"],
                *[r["jumlah"][t] for t, _lbl in _SDM_KATEGORI_TABEL],
                r["total_entri"],
                r["kesiapan_bkd"],
                r["kesiapan_sister"],
                r["n_reminder"],
            )
        )
    return _kirim_excel("Rekap Kinerja Dosen", headers, data)


# ---------------------------------------------------- Rekap Program Kerja & Kegiatan
# Bab Kegiatan/Program Kerja Prodi di Rekap & Laporan — sama seperti bab SDM
# di atas, memakai rumus realisasi yang SAMA PERSIS dgn kegiatan.py
# (_hitung_realisasi: rasio kegiatan berstatus 'Selesai') supaya konsisten
# dgn yang tampil di halaman Program Kerja, ditambah agregat per bidang +
# ekspor Excel yang belum ada di sana.
def _rows_program_kerja(conn, tahun_akademik=None):
    from app.routes.kegiatan import _hitung_realisasi

    q = "SELECT * FROM program_kerja"
    params = []
    if tahun_akademik:
        q += " WHERE tahun_akademik=?"
        params.append(tahun_akademik)
    q += " ORDER BY tahun_akademik DESC, bidang, nama_program"
    proker_rows = conn.execute(q, params).fetchall()
    detail = []
    for p in proker_rows:
        total, selesai, persen = _hitung_realisasi(conn, p["id"])
        detail.append({"row": p, "total": total, "selesai": selesai, "persen": persen})

    per_bidang = {}
    for d in detail:
        b = d["row"]["bidang"] or "Lainnya"
        agg = per_bidang.setdefault(
            b,
            {
                "jumlah_program": 0,
                "jumlah_kegiatan": 0,
                "jumlah_selesai": 0,
                "anggaran_rencana": 0.0,
            },
        )
        agg["jumlah_program"] += 1
        agg["jumlah_kegiatan"] += d["total"]
        agg["jumlah_selesai"] += d["selesai"]
        agg["anggaran_rencana"] += d["row"]["anggaran_rencana"] or 0
    ringkasan_bidang = []
    for bidang, agg in sorted(per_bidang.items()):
        persen = (
            round(100 * agg["jumlah_selesai"] / agg["jumlah_kegiatan"])
            if agg["jumlah_kegiatan"]
            else 0
        )
        ringkasan_bidang.append({"bidang": bidang, "persen": persen, **agg})
    return detail, ringkasan_bidang


@bp.route("/program-kerja")
def program_kerja():
    conn = current_app.get_db()
    tahun_akademik = request.args.get("tahun_akademik", "") or None
    detail, ringkasan_bidang = _rows_program_kerja(conn, tahun_akademik)
    tahun_list = [
        r["tahun_akademik"]
        for r in conn.execute(
            "SELECT DISTINCT tahun_akademik FROM program_kerja WHERE tahun_akademik IS NOT NULL "
            "AND tahun_akademik != '' ORDER BY tahun_akademik DESC"
        ).fetchall()
    ]
    return render_template(
        "rekap_program_kerja.html",
        detail=detail,
        ringkasan_bidang=ringkasan_bidang,
        tahun_akademik=tahun_akademik or "",
        tahun_list=tahun_list,
    )


@bp.route("/program-kerja/ekspor")
def program_kerja_ekspor():
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    conn = current_app.get_db()
    tahun_akademik = request.args.get("tahun_akademik", "") or None
    detail, ringkasan_bidang = _rows_program_kerja(conn, tahun_akademik)
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Bagian 1 - Detail Program"
    ws1.append(
        [
            "Tahun Akademik",
            "Bidang",
            "Nama Program",
            "Penanggung Jawab",
            "Status",
            "Anggaran Rencana",
            "Jumlah Kegiatan",
            "Kegiatan Selesai",
            "Realisasi (%)",
        ]
    )
    for d in detail:
        p = d["row"]
        ws1.append(
            [
                p["tahun_akademik"],
                p["bidang"],
                p["nama_program"],
                p["penanggung_jawab"],
                p["status"],
                p["anggaran_rencana"],
                d["total"],
                d["selesai"],
                d["persen"],
            ]
        )
    ws2 = wb.create_sheet("Bagian 2 - Ringkasan Bidang")
    ws2.append(
        [
            "Bidang",
            "Jumlah Program",
            "Jumlah Kegiatan",
            "Kegiatan Selesai",
            "Anggaran Rencana",
            "Realisasi (%)",
        ]
    )
    for r in ringkasan_bidang:
        ws2.append(
            [
                r["bidang"],
                r["jumlah_program"],
                r["jumlah_kegiatan"],
                r["jumlah_selesai"],
                r["anggaran_rencana"],
                r["persen"],
            ]
        )
    for ws in (ws1, ws2):
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1E3A5F")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="Rekap_Program_Kerja.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ------------------------------------------------------------------- Statistik
@bp.route("/statistik")
def statistik():
    conn = current_app.get_db()
    counts = L.dashboard_counts(conn)
    from app.constants import STATUS_TA_LIST

    chart = [(s, counts.get(s, 0)) for s in STATUS_TA_LIST]
    max_val = max([c for _, c in chart] + [1])
    return render_template("statistik.html", counts=counts, chart=chart, max_val=max_val)
