# -*- coding: utf-8 -*-
"""routes/pengaturan.py — Manajemen Pengguna, Identitas & Branding,
Import & Export Data, Tahun Akademik, Ubah Password."""

import io
import os

from flask import (
    Blueprint,
    current_app,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from werkzeug.utils import secure_filename

from app import auth_core
from app import constants as C
from app import db as _db
from app import error_utils as EH
from app import import_excel, import_generic
from app.pin_guard import perlu_pin, tandai_pin_terverifikasi

bp = Blueprint("pengaturan", __name__, url_prefix="/pengaturan")


# ---------------------------------------------------------------- Pengguna
@bp.route("/pengguna")
def pengguna():
    conn = current_app.get_db()
    rows = conn.execute("SELECT * FROM pengguna ORDER BY id").fetchall()
    return render_template("pengaturan/pengguna.html", rows=rows)


@bp.route("/pengguna/tambah", methods=["POST"])
def pengguna_tambah():
    conn = current_app.get_db()
    nama = request.form.get("nama", "").strip()
    peran = request.form.get("peran", "Administrator")
    if nama:
        conn.execute("INSERT INTO pengguna(nama, peran, aktif) VALUES(?,?,1)", (nama, peran))
        conn.commit()
        flash(f"Pengguna {nama} ditambahkan.", "ok")
    return redirect(url_for("pengaturan.pengguna"))


# ------------------------------------------------------- Identitas & Branding
BIDANG_BRANDING = [
    ("nama_institusi", "Nama Kampus / Institusi"),
    ("nama_fakultas", "Nama Fakultas"),
    ("nama_prodi", "Nama Program Studi"),
    ("email", "Email Resmi"),
    ("telp", "Telepon"),
    ("alamat", "Alamat"),
]

UPLOAD_EXT = {"png", "jpg", "jpeg", "svg"}


@bp.route("/branding", methods=["GET", "POST"])
def branding():
    conn = current_app.get_db()
    if request.method == "POST":
        for key, _label in BIDANG_BRANDING:
            _db.set_setting(conn, key, request.form.get(key, "").strip())

        file = request.files.get("logo")
        if file and file.filename:
            ext = file.filename.rsplit(".", 1)[-1].lower()
            if ext in UPLOAD_EXT:
                folder = os.path.join(_db.home_dir(), "SistemSkripsi", "branding")
                os.makedirs(folder, exist_ok=True)
                fname = secure_filename(f"logo.{ext}")
                dest = os.path.join(folder, fname)
                file.save(dest)
                _db.set_setting(conn, "logo_path", dest)
            else:
                flash("Format logo harus PNG, JPG, atau SVG.", "error")

        _db.log(conn, "Ubah Identitas & Branding")
        flash("Identitas & branding tersimpan.", "ok")
        return redirect(url_for("pengaturan.branding"))

    nilai = {key: _db.get_setting(conn, key, "") for key, _ in BIDANG_BRANDING}
    return render_template("pengaturan/branding.html", nilai=nilai, bidang=BIDANG_BRANDING)


@bp.route("/pejabat")
def pejabat():
    """Direktori Pejabat Struktural (Rektor/Dekan/Kaprodi/dst). Lihat
    komentar CREATE TABLE pejabat_struktural di db.py untuk latar
    belakang — tabel ini menggantikan 2 kolom teks bebas
    (nama_penandatangan_default/jabatan_penandatangan_default) sebagai
    sumber data terstruktur, dengan setting lama tetap disinkron sebagai
    cache tampilan supaya modul yang belum diarahkan ke sini (SK Tugas
    Akhir/Surat Umum) tetap dapat nama penandatangan aktif."""
    conn = current_app.get_db()
    rows = conn.execute(
        "SELECT * FROM pejabat_struktural ORDER BY urutan, jabatan, nama"
    ).fetchall()
    edit_id = request.args.get("edit", type=int)
    edit_row = None
    if edit_id:
        edit_row = conn.execute(
            "SELECT * FROM pejabat_struktural WHERE id=?", (edit_id,)
        ).fetchone()
    default_saat_ini = _db.get_setting(conn, "nama_penandatangan_default", "")
    return render_template(
        "pengaturan/pejabat.html",
        rows=rows,
        edit_row=edit_row,
        default_saat_ini=default_saat_ini,
        DAFTAR_JABATAN_STRUKTURAL=C.DAFTAR_JABATAN_STRUKTURAL,
    )


@bp.route("/pejabat/simpan", methods=["POST"])
def pejabat_simpan():
    conn = current_app.get_db()
    f = request.form
    pid = f.get("id", type=int)
    jabatan = f.get("jabatan", "").strip()
    nama = f.get("nama", "").strip()
    if not jabatan or not nama:
        flash("Jabatan dan Nama wajib diisi.", "error")
        return redirect(url_for("pengaturan.pejabat"))
    data = (
        jabatan,
        f.get("unit", "").strip(),
        nama,
        f.get("nip_nidn", "").strip(),
        f.get("no_sk_pengangkatan", "").strip(),
        f.get("tmt", "").strip(),
        f.get("masa_akhir", "").strip(),
        f.get("urutan", type=int) or 0,
        1 if f.get("aktif") == "on" else 0,
    )
    try:
        if pid:
            conn.execute(
                "UPDATE pejabat_struktural SET jabatan=?, unit=?, nama=?, nip_nidn=?, "
                "no_sk_pengangkatan=?, tmt=?, masa_akhir=?, urutan=?, aktif=? WHERE id=?",
                data + (pid,),
            )
            _db.log(conn, "Update Pejabat Struktural", f"{jabatan} — {nama}")
            flash("Data pejabat diperbarui.", "ok")
        else:
            conn.execute(
                "INSERT INTO pejabat_struktural(jabatan,unit,nama,nip_nidn,"
                "no_sk_pengangkatan,tmt,masa_akhir,urutan,aktif) VALUES(?,?,?,?,?,?,?,?,?)",
                data,
            )
            _db.log(conn, "Tambah Pejabat Struktural", f"{jabatan} — {nama}")
            flash(f"Pejabat {nama} ({jabatan}) ditambahkan.", "ok")
        conn.commit()
    except Exception as e:
        EH.flash_gagal_simpan(e, "Gagal menyimpan data pejabat")
    return redirect(url_for("pengaturan.pejabat"))


@bp.route("/pejabat/<int:pid>/hapus", methods=["POST"])
def pejabat_hapus(pid):
    conn = current_app.get_db()
    row = conn.execute("SELECT nama FROM pejabat_struktural WHERE id=?", (pid,)).fetchone()
    conn.execute("DELETE FROM pejabat_struktural WHERE id=?", (pid,))
    conn.commit()
    if row:
        _db.log(conn, "Hapus Pejabat Struktural", row["nama"])
    flash("Data pejabat dihapus.", "ok")
    return redirect(url_for("pengaturan.pejabat"))


@bp.route("/pejabat/<int:pid>/jadikan-default", methods=["POST"])
def pejabat_jadikan_default(pid):
    """Menjadikan 1 pejabat sebagai penandatangan default SK Tugas Akhir
    (routes/surat.py _footer_ttd) & Surat Umum (routes/surat_umum.py) —
    disinkron ke 3 kunci `pengaturan` lama (nama/jabatan/nip_nidn
    penandatangan default) supaya kedua modul itu tidak perlu diubah lagi
    tiap kali direktori pejabat berubah."""
    conn = current_app.get_db()
    row = conn.execute("SELECT * FROM pejabat_struktural WHERE id=?", (pid,)).fetchone()
    if not row:
        flash("Data pejabat tidak ditemukan.", "error")
        return redirect(url_for("pengaturan.pejabat"))
    _db.set_setting(conn, "nama_penandatangan_default", row["nama"])
    _db.set_setting(conn, "jabatan_penandatangan_default", row["jabatan"])
    _db.set_setting(conn, "nip_nidn_penandatangan_default", row["nip_nidn"] or "")
    _db.log(conn, "Set Penandatangan Default", f"{row['jabatan']} — {row['nama']}")
    flash(
        f'"{row["nama"]}" ({row["jabatan"]}) dijadikan penandatangan default untuk '
        "SK Tugas Akhir & Surat Umum.",
        "ok",
    )
    return redirect(url_for("pengaturan.pejabat"))


@bp.route("/logo-preview")
def logo_preview():
    conn = current_app.get_db()
    path = _db.get_setting(conn, "logo_path", "")
    if not path or not os.path.exists(path):
        from flask import abort

        abort(404)
    return send_file(path)


# --------------------------------------------------------------- Tarif Honor
# Audit Modul Pelaksanaan — tarif_honor_seminar/tarif_honor_penguji_sidang/
# tarif_honor_pembimbing_1/2 sudah ada sebagai `pengaturan` (dipakai
# logic.rkp_seminar/rkp_sidang & snapshot honor di routes/pelaksanaan.py),
# tapi sebelum ini TIDAK ADA halaman UI untuk mengubahnya — cuma bisa
# diedit langsung di file database. Halaman ini melengkapi itu.
BIDANG_HONOR = [
    ("tarif_honor_seminar", "Honor Penguji Seminar (per peran, per mahasiswa)"),
    ("tarif_honor_penguji_sidang", "Honor Penguji Sidang (per peran, per mahasiswa)"),
    ("tarif_honor_pembimbing_1", "Honor Pembimbing 1 (per mahasiswa LULUS sidang)"),
    ("tarif_honor_pembimbing_2", "Honor Pembimbing 2 (per mahasiswa LULUS sidang)"),
]


@bp.route("/honor", methods=["GET", "POST"])
def honor():
    conn = current_app.get_db()
    if request.method == "POST":
        for key, _label in BIDANG_HONOR:
            mentah = request.form.get(key, "").strip().replace(".", "").replace(",", "")
            if mentah and not mentah.isdigit():
                flash(f"Tarif untuk \"{dict(BIDANG_HONOR)[key]}\" harus berupa angka.", "error")
                return redirect(url_for("pengaturan.honor"))
            _db.set_setting(conn, key, mentah or "0")
        _db.log(conn, "Ubah Tarif Honor")
        flash(
            "Tarif honor tersimpan. Perubahan ini HANYA berlaku untuk honor baru yang "
            "terbentuk setelah ini (status seminar diset 'Selesai' / sidang diset 'LULUS') "
            "— honor tahap-tahap sebelumnya tidak berubah retroaktif.",
            "ok",
        )
        return redirect(url_for("pengaturan.honor"))

    nilai = {key: _db.get_setting(conn, key, "0") for key, _ in BIDANG_HONOR}
    return render_template("pengaturan/honor.html", nilai=nilai, bidang=BIDANG_HONOR)


# --------------------------------------------------------------- Tahun Akademik
@bp.route("/tahun-akademik", methods=["GET", "POST"])
def tahun_akademik():
    """Wizard "Buka Tahun Ajaran Baru" (Audit poin 1, rekomendasi #5) —
    menggantikan 3 text-input bebas (tahun_akademik_aktif, nama_tahap_1/2)
    dengan struktur relasional tahun_ajaran -> periode_akademik ->
    tahap_pengajuan. Pengaturan lama (kunci `pengaturan`) tetap disinkron
    sebagai cache tampilan (dipakai context_processor & modul yang belum
    dimigrasikan penuh), jadi tidak ada modul lama yang tiba-tiba kosong."""
    conn = current_app.get_db()

    if request.method == "POST":
        aksi = request.form.get("aksi", "buka_tahun")
        if aksi == "buka_tahun":
            kode = request.form.get("kode", "").strip()
            aktifkan = request.form.get("aktifkan", "Ganjil")
            try:
                _db.buka_tahun_ajaran(conn, kode, aktifkan=aktifkan)
                _db.log(conn, "Buka Tahun Ajaran", kode)
                flash(f"Tahun ajaran {kode} dibuka (Ganjil/Genap/Antara sekaligus dibuat).", "ok")
            except ValueError as e:
                flash(str(e), "error")
        elif aksi == "ubah_kode_ta":
            # Restrukturisasi poin 3 — perbaikan salah ketik kode tahun
            # ajaran (mis. "2025/2025" -> "2025/2026"). SENGAJA tidak ada
            # aksi "hapus_tahun_ajaran"/"hapus_periode" di route ini —
            # lihat db.ubah_kode_tahun_ajaran() untuk alasan kenapa edit
            # aman dilakukan kapan pun (relasi ke tahun ajaran memakai id,
            # bukan kode) sementara hapus sengaja tidak disediakan.
            ta_id = request.form.get("ta_id", type=int)
            kode_baru = request.form.get("kode_baru", "").strip()
            ok, pesan = _db.ubah_kode_tahun_ajaran(conn, ta_id, kode_baru)
            if ok:
                _db.log(conn, "Ubah Kode Tahun Ajaran", f"#{ta_id} -> {kode_baru}")
                flash(f'Kode tahun ajaran diperbarui menjadi "{kode_baru}".', "ok")
            else:
                flash(pesan, "error")
        elif aksi == "set_status_periode":
            periode_id = request.form.get("periode_id", type=int)
            status = request.form.get("status", "Draft")
            if status not in C.STATUS_PERIODE_LIST:
                status = "Draft"
            periode = conn.execute(
                "SELECT * FROM periode_akademik WHERE id=?", (periode_id,)
            ).fetchone()
            if periode:
                if status == "Berjalan":
                    # Hanya 1 periode boleh 'Berjalan' pada satu waktu, supaya
                    # get_periode_aktif() (dipakai form dropdown semester di
                    # seluruh modul) tidak ambigu.
                    conn.execute(
                        "UPDATE periode_akademik SET status='Draft' WHERE status='Berjalan'"
                    )
                    ta = conn.execute(
                        "SELECT kode FROM tahun_ajaran WHERE id=?", (periode["tahun_ajaran_id"],)
                    ).fetchone()
                    if ta:
                        _db.set_setting(conn, "tahun_akademik_aktif", ta["kode"])
                conn.execute(
                    "UPDATE periode_akademik SET status=? WHERE id=?", (status, periode_id)
                )
                conn.commit()
                flash("Status periode diperbarui.", "ok")
        elif aksi == "tambah_tahap":
            periode_id = request.form.get("periode_id", type=int)
            nama = request.form.get("nama_tahap", "").strip()
            if not periode_id or not nama:
                flash("Pilih periode & isi nama tahap.", "error")
            else:
                urutan = conn.execute(
                    "SELECT COALESCE(MAX(urutan),0)+1 u FROM tahap_pengajuan WHERE periode_akademik_id=?",
                    (periode_id,),
                ).fetchone()["u"]
                conn.execute(
                    "INSERT INTO tahap_pengajuan(periode_akademik_id, urutan, nama, tgl_buka, tgl_tutup) "
                    "VALUES(?,?,?,?,?)",
                    (
                        periode_id,
                        urutan,
                        nama,
                        request.form.get("tgl_buka", ""),
                        request.form.get("tgl_tutup", ""),
                    ),
                )
                conn.commit()
                _db.log(conn, "Tambah Tahap Pengajuan", nama)
                flash(f'Tahap "{nama}" ditambahkan.', "ok")
        elif aksi == "hapus_tahap":
            # Audit Kontinuitas: pengajuan_judul.tahap_pengajuan_id,
            # penetapan_pembimbing.tahap_pengajuan_id, seminar.tahap_pengajuan_id,
            # dan sidang.tahap_pengajuan_id ditambahkan lewat ALTER TABLE tanpa
            # constraint FK sungguhan (keterbatasan SQLite), jadi database
            # tidak akan mencegah sendiri penghapusan tahap yang masih dipakai
            # mahasiswa aktif. Tanpa guard ini, baris mahasiswa yang
            # mereferensikan tahap yang dihapus jadi "orphan" — datanya tidak
            # hilang, tapi diam-diam tidak muncul lagi di Rekap Honor/SK
            # Yudisium per tahap karena join ke tahap_pengajuan gagal.
            tahap_id = request.form.get("tahap_id", type=int)
            tahap = conn.execute(
                "SELECT nama FROM tahap_pengajuan WHERE id=?", (tahap_id,)
            ).fetchone()
            if not tahap:
                flash("Tahap tidak ditemukan.", "error")
                return redirect(url_for("pengaturan.tahun_akademik"))
            pemakaian = {
                "Pengajuan Judul": "pengajuan_judul",
                "Penetapan Pembimbing": "penetapan_pembimbing",
                "Seminar": "seminar",
                "Sidang": "sidang",
            }
            rincian = []
            for label, tabel in pemakaian.items():
                n = conn.execute(
                    f"SELECT COUNT(*) n FROM {tabel} WHERE tahap_pengajuan_id=?",
                    (tahap_id,),
                ).fetchone()["n"]
                if n:
                    rincian.append(f"{label} ({n})")
            if rincian:
                flash(
                    f"Tahap \"{tahap['nama']}\" tidak bisa dihapus — masih "
                    "dirujuk oleh data mahasiswa di: " + ", ".join(rincian) +
                    ". Menghapusnya akan membuat data tersebut hilang dari "
                    "Rekap Honor & SK Yudisium per tahap. Pindahkan data "
                    "mahasiswa ke tahap lain dulu, baru tahap ini bisa dihapus.",
                    "error",
                )
                return redirect(url_for("pengaturan.tahun_akademik"))
            conn.execute("DELETE FROM tahap_pengajuan WHERE id=?", (tahap_id,))
            conn.commit()
            _db.log(conn, "Hapus Tahap Pengajuan", tahap["nama"])
            flash("Tahap dihapus.", "ok")
        return redirect(url_for("pengaturan.tahun_akademik"))

    tahun_ajaran_list = conn.execute("SELECT * FROM tahun_ajaran ORDER BY id DESC").fetchall()
    periode_by_ta = {}
    tahap_by_periode = {}
    for ta in tahun_ajaran_list:
        periodes = conn.execute(
            "SELECT * FROM periode_akademik WHERE tahun_ajaran_id=? ORDER BY "
            "CASE jenis WHEN 'Ganjil' THEN 1 WHEN 'Genap' THEN 2 ELSE 3 END",
            (ta["id"],),
        ).fetchall()
        periode_by_ta[ta["id"]] = periodes
        for p in periodes:
            tahap_by_periode[p["id"]] = conn.execute(
                "SELECT * FROM tahap_pengajuan WHERE periode_akademik_id=? ORDER BY urutan",
                (p["id"],),
            ).fetchall()
    periode_aktif = _db.get_periode_aktif(conn)
    return render_template(
        "pengaturan/tahun_akademik.html",
        tahun_ajaran_list=tahun_ajaran_list,
        periode_by_ta=periode_by_ta,
        tahap_by_periode=tahap_by_periode,
        periode_aktif=periode_aktif,
        JENIS_PERIODE_LIST=C.JENIS_PERIODE_LIST,
        STATUS_PERIODE_LIST=C.STATUS_PERIODE_LIST,
        tahun_akademik_aktif_lama=_db.get_setting(conn, "tahun_akademik_aktif", ""),
    )


# --------------------------------------------------------------- Ubah Username & Password
@bp.route("/password", methods=["GET", "POST"])
def password():
    """Restrukturisasi poin 1 — username sekarang bisa diganti di sini
    juga (dulu hanya password). Ganti password tetap wajib mengonfirmasi
    password LAMA (tidak berubah dari perilaku semula); ganti username
    tidak minta konfirmasi ulang karena username bukan rahasia — hanya
    identitas tampilan."""
    conn = current_app.get_db()
    if request.method == "POST":
        aksi = request.form.get("aksi", "ubah_password")
        if aksi == "ubah_username":
            username_baru = request.form.get("username_baru", "").strip()
            if not username_baru:
                flash("Username baru wajib diisi.", "error")
            else:
                username_lama = auth_core.get_username(conn)
                auth_core.set_username(conn, username_baru)
                _db.log(conn, "Ubah Username", f"{username_lama} -> {username_baru}")
                flash("Username berhasil diubah.", "ok")
            return redirect(url_for("pengaturan.password"))
        else:
            lama = request.form.get("lama", "")
            baru1 = request.form.get("baru1", "")
            baru2 = request.form.get("baru2", "")
            if not auth_core.verify_password(conn, lama):
                flash("Password lama salah.", "error")
            elif len(baru1) < 4:
                flash("Password baru minimal 4 karakter.", "error")
            elif baru1 != baru2:
                flash("Konfirmasi password baru tidak sama.", "error")
            else:
                auth_core.set_password(conn, baru1)
                _db.log(conn, "Ubah Password")
                flash("Password berhasil diubah.", "ok")
            return redirect(url_for("pengaturan.password"))
    return render_template("pengaturan/password.html", username_sekarang=auth_core.get_username(conn))


# --------------------------------------------------------------- PIN Fitur Krusial
@bp.route("/pin", methods=["GET", "POST"])
def pin_atur():
    """Restrukturisasi poin 2 — atur/ganti PIN tambahan untuk membuka
    fitur krusial (Import Data, Restore Backup, dst — lihat
    app/pin_guard.py). PIN sengaja divalidasi harus BERBEDA dari password
    login, supaya benar-benar jadi lapis proteksi kedua yang independen,
    bukan sekadar password yang ditulis ulang."""
    conn = current_app.get_db()
    sudah_ada = auth_core.has_pin(conn)
    wajib = request.args.get("wajib") == "1"
    if request.method == "POST":
        pin_lama = request.form.get("pin_lama", "")
        pin1 = request.form.get("pin1", "").strip()
        pin2 = request.form.get("pin2", "").strip()
        if sudah_ada and not auth_core.verify_pin(conn, pin_lama):
            flash("PIN lama salah.", "error")
        elif not (pin1.isdigit() and 4 <= len(pin1) <= 8):
            flash("PIN harus 4-8 digit angka.", "error")
        elif pin1 != pin2:
            flash("Konfirmasi PIN tidak sama.", "error")
        elif auth_core.verify_password(conn, pin1):
            flash("PIN tidak boleh sama dengan password login — gunakan kombinasi lain.", "error")
        else:
            auth_core.set_pin(conn, pin1)
            auth_core.reset_percobaan_gagal(conn, prefix="pin")
            _db.log(conn, "Atur PIN Fitur Krusial")
            flash("PIN berhasil disimpan.", "ok")
            tujuan = session.pop("pin_tujuan", None)
            tandai_pin_terverifikasi()
            return redirect(tujuan or url_for("pengaturan.pin_atur"))
    return render_template("pengaturan/pin_atur.html", sudah_ada=sudah_ada, wajib=wajib)


@bp.route("/pin/verifikasi", methods=["GET", "POST"])
def pin_verifikasi():
    """Halaman jembatan sebelum masuk ke menu berpagar PIN (dilempar ke
    sini oleh @perlu_pin). Memakai mekanisme lockout yang sama dengan
    login (prefix="pin" — terpisah dari percobaan gagal login itu
    sendiri)."""
    conn = current_app.get_db()
    if not auth_core.has_pin(conn):
        return redirect(url_for("pengaturan.pin_atur", wajib=1))

    lockout_sisa = auth_core.cek_lockout(conn, prefix="pin")
    if request.method == "POST" and lockout_sisa:
        flash(f"Terlalu banyak percobaan gagal. Coba lagi dalam {lockout_sisa // 60 + 1} menit.", "error")
    elif request.method == "POST":
        pin = request.form.get("pin", "")
        if auth_core.verify_pin(conn, pin):
            auth_core.reset_percobaan_gagal(conn, prefix="pin")
            tandai_pin_terverifikasi()
            tujuan = session.pop("pin_tujuan", None)
            return redirect(tujuan or url_for("dashboard.index"))
        else:
            auth_core.catat_percobaan_gagal(conn, prefix="pin")
            sisa = auth_core.cek_lockout(conn, prefix="pin")
            if sisa:
                flash(f"Terlalu banyak percobaan gagal. Verifikasi PIN dikunci {sisa // 60 + 1} menit.", "error")
            else:
                flash("PIN salah.", "error")
    return render_template("pengaturan/pin_verifikasi.html", lockout_sisa=lockout_sisa)


# ------------------------------------------------------- Import & Export Data
# Restrukturisasi poin 2 — @perlu_pin dipasang di SEMUA route Import (baik
# halaman migrasi 1x/import_export, maupun Import Generik rutin di bawah)
# supaya menu ini benar-benar tidak bisa dibuka tanpa PIN, persis contoh
# yang diminta ("ketika akan membuka menu import perlu memasukan pin
# terlebih dahulu").
@bp.route("/import-export", methods=["GET", "POST"])
@perlu_pin
def import_export():
    conn = current_app.get_db()
    ringkasan = None
    if request.method == "POST":
        file = request.files.get("file_excel")
        # Audit Menyeluruh — migrasi data produksi nyata (Agustus 2026):
        # workbook "Aplikasi Manajemen RPL" institusi ini strukturnya
        # BERBEDA TOTAL dari workbook skripsi reguler (sheet tergabung,
        # nama kolom beda total) -- run_import() lama TIDAK BISA membaca
        # workbook RPL sama sekali (nama sheet yang dicarinya tidak akan
        # pernah ketemu, hasilnya diam-diam 0 baris diimpor tanpa
        # peringatan). jenis_workbook menentukan fungsi importer mana yang
        # dipanggil -- BUKAN sekadar skema_default seperti sebelumnya
        # (dropdown itu tetap ada, tapi kini HANYA relevan utk workbook
        # Reguler; workbook RPL selalu memberi skema='RPL' tanpa opsi).
        jenis_workbook = request.form.get("jenis_workbook", "reguler")
        skema_default = request.form.get("skema_default", "Reguler")
        if not file or not file.filename:
            flash("Pilih file Excel (.xlsx) terlebih dahulu.", "error")
        elif not file.filename.lower().endswith(".xlsx"):
            flash("File harus berformat .xlsx.", "error")
        else:
            tmp_dir = os.path.join(_db.home_dir(), "SistemSkripsi", "tmp_import")
            os.makedirs(tmp_dir, exist_ok=True)
            tmp_path = os.path.join(tmp_dir, secure_filename(file.filename))
            file.save(tmp_path)
            try:
                if jenis_workbook == "rpl":
                    ringkasan = import_excel.run_import_rpl(tmp_path, conn)
                    _db.log(conn, "Import Excel RPL", file.filename)
                else:
                    ringkasan = import_excel.run_import(tmp_path, conn, skema_default=skema_default)
                    _db.log(conn, "Import Excel", file.filename)
                flash("Import selesai — lihat ringkasan di bawah.", "ok")
            except Exception as e:
                EH.flash_gagal_simpan(e, "Import gagal")
            finally:
                try:
                    os.remove(tmp_path)
                except OSError:
                    pass
    return render_template("pengaturan/import_export.html", ringkasan=ringkasan)


# --------------------------------------------------- Import Generik per-Modul
@bp.route("/import-generik", methods=["GET"])
@perlu_pin
def import_generik():
    """Import Data Generik (Audit poin 6) — dipakai rutin, beda dari
    /import-export (migrasi 1x dari workbook lama lengkap di atas).
    Dimulai dari Dosen & Mahasiswa (Reguler & RPL) sesuai prioritas."""
    modul = request.args.get("modul", "dosen")
    if modul not in import_generic.IMPORTERS:
        modul = "dosen"
    return render_template(
        "pengaturan/import_generik.html",
        modul=modul,
        importers=import_generic.IMPORTERS,
        hasil=None,
    )


#: Karakter yang tidak boleh ada di nama file Windows (juga bermasalah
#: kalau lolos ke header Content-Disposition, mis. "/" bisa kebaca browser
#: sebagai pemisah folder). Beberapa label modul (mis. "SDM — Luaran
#: (Publikasi/HKI/Buku/dst)") sebelumnya mengandung "/" di sini,
#: menghasilkan nama unduhan yang rusak; sekarang disaring dulu.
_INVALID_FILENAME_CHARS = set('\\/:*?"<>|')


def _nama_file_aman(label):
    bersih = "".join(ch for ch in label if ch not in _INVALID_FILENAME_CHARS)
    return bersih.replace(" ", "_")


@bp.route("/import-generik/template/<modul>")
def import_generik_template(modul):
    if modul not in import_generic.IMPORTERS:
        flash("Modul import tidak dikenal.", "error")
        return redirect(url_for("pengaturan.import_generik"))
    buf = import_generic.buat_template(modul)
    label = import_generic.IMPORTERS[modul]["label"]
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"Template_{_nama_file_aman(label)}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.route("/import-generik/proses", methods=["POST"])
@perlu_pin
def import_generik_proses():
    conn = current_app.get_db()
    modul = request.form.get("modul", "dosen")
    if modul not in import_generic.IMPORTERS:
        flash("Modul import tidak dikenal.", "error")
        return redirect(url_for("pengaturan.import_generik"))
    file = request.files.get("file_excel")
    if not file or not file.filename:
        flash("Pilih file hasil isian template terlebih dahulu.", "error")
        return redirect(url_for("pengaturan.import_generik", modul=modul))
    if not file.filename.lower().endswith(".xlsx"):
        flash("File harus berformat .xlsx (hasil dari template yang diunduh).", "error")
        return redirect(url_for("pengaturan.import_generik", modul=modul))
    hasil = import_generic.proses_upload(modul, file.stream, conn)
    if "error" in hasil:
        flash(hasil["error"], "error")
        return redirect(url_for("pengaturan.import_generik", modul=modul))
    _db.log(
        conn,
        f"Import Generik: {import_generic.IMPORTERS[modul]['label']}",
        f"{hasil['tambah']} baru, {hasil['update']} diperbarui, {hasil['lewati']} dilewati",
    )
    flash(
        f"Selesai — {hasil['tambah']} data baru, {hasil['update']} diperbarui, "
        f"{hasil['lewati']} dilewati.",
        "ok",
    )
    return render_template(
        "pengaturan/import_generik.html",
        modul=modul,
        importers=import_generic.IMPORTERS,
        hasil=hasil,
    )


def _export_generic(sheet_title, headers, rows):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1E3A5F")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for r in rows:
        ws.append(list(r))
    for col in ws.columns:
        length = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(length + 2, 10), 45)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@bp.route("/export/dosen")
def export_dosen():
    conn = current_app.get_db()
    rows = conn.execute(
        "SELECT nidn, nik, nuptk, nama, no_hp, email FROM dosen ORDER BY nama"
    ).fetchall()
    buf = _export_generic(
        "Data Dosen",
        ["NIDN", "NIK", "NUPTK", "Nama", "No HP", "Email"],
        [(r["nidn"], r["nik"], r["nuptk"], r["nama"], r["no_hp"], r["email"]) for r in rows],
    )
    return send_file(
        buf,
        as_attachment=True,
        download_name="Data_Dosen.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
