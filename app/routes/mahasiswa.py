# -*- coding: utf-8 -*-
"""routes/mahasiswa.py — CRUD Data Mahasiswa, termasuk filter Skema
(Reguler/RPL) dan ekspor Excel."""

import io

from flask import (
    Blueprint,
    current_app,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)

from app import constants as C
from app import db as _db
from app import error_utils as EH

bp = Blueprint("mahasiswa", __name__, url_prefix="/mahasiswa")


@bp.route("/")
def list_view():
    conn = current_app.get_db()
    q = request.args.get("q", "").strip()
    skema = request.args.get("skema", "")

    sql = "SELECT * FROM mahasiswa WHERE 1=1"
    params = []
    if q:
        sql += " AND (nim LIKE ? OR nama LIKE ?)"
        params += [f"%{q}%", f"%{q}%"]
    if skema in ("Reguler", "RPL"):
        sql += " AND skema=?"
        params.append(skema)
    sql += " ORDER BY nama"
    rows = conn.execute(sql, params).fetchall()

    total = conn.execute("SELECT COUNT(*) c FROM mahasiswa").fetchone()["c"]
    n_reguler = conn.execute(
        "SELECT COUNT(*) c FROM mahasiswa WHERE skema IS NULL OR skema='Reguler'"
    ).fetchone()["c"]
    n_rpl = conn.execute("SELECT COUNT(*) c FROM mahasiswa WHERE skema='RPL'").fetchone()["c"]

    return render_template(
        "mahasiswa_list.html",
        rows=rows,
        q=q,
        skema=skema,
        total=total,
        n_reguler=n_reguler,
        n_rpl=n_rpl,
        STATUS_TA_LIST=C.STATUS_TA_LIST,
        STATUS_MHS_LIST=C.STATUS_MHS_LIST,
    )


@bp.route("/tambah", methods=["GET", "POST"])
def tambah():
    conn = current_app.get_db()
    if request.method == "POST":
        f = request.form
        nim = f.get("nim", "").strip()
        nama = f.get("nama", "").strip()
        if not nim or not nama:
            flash("NIM dan Nama wajib diisi.", "error")
        else:
            try:
                conn.execute(
                    "INSERT INTO mahasiswa(nim,nama,jk,tempat_lahir,tgl_lahir,no_hp,email_nik,"
                    "angkatan,konsentrasi,status,status_ta,skema,catatan) "
                    "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)",
                    (
                        nim,
                        nama,
                        f.get("jk", ""),
                        f.get("tempat_lahir", ""),
                        f.get("tgl_lahir", ""),
                        f.get("no_hp", ""),
                        f.get("email_nik", ""),
                        f.get("angkatan", ""),
                        f.get("konsentrasi", ""),
                        f.get("status", "Aktif"),
                        # Audit P0 #2 — status_ta TIDAK LAGI diambil dari form.
                        # Mahasiswa baru SELALU mulai dari "Belum Mengajukan
                        # Judul"; status_ta sejak itu HANYA berubah lewat
                        # logic.recalculate_status_ta() (dipicu oleh histori
                        # pengajuan/pembimbing/seminar/sidang yang sesungguhnya
                        # -- lihat routes/akademik.py & routes/pelaksanaan.py),
                        # bukan lagi field yang bisa ditimpa manual di sini.
                        C.STATUS_TA_BELUM,
                        f.get("skema", "Reguler"),
                        f.get("catatan", ""),
                    ),
                )
                conn.commit()
                _db.log(conn, "Tambah Mahasiswa", f"{nim} - {nama}")
                flash(f"Mahasiswa {nama} berhasil ditambahkan.", "ok")
                return redirect(url_for("mahasiswa.list_view"))
            except Exception as e:
                EH.flash_gagal_simpan(e, "Gagal menyimpan")
    return render_template(
        "mahasiswa_form.html",
        row=None,
        STATUS_TA_LIST=C.STATUS_TA_LIST,
        STATUS_MHS_LIST=C.STATUS_MHS_LIST,
        JK_LIST=C.JK_LIST,
    )


@bp.route("/<int:mid>/edit", methods=["GET", "POST"])
def edit(mid):
    conn = current_app.get_db()
    row = conn.execute("SELECT * FROM mahasiswa WHERE id=?", (mid,)).fetchone()
    if not row:
        flash("Data mahasiswa tidak ditemukan.", "error")
        return redirect(url_for("mahasiswa.list_view"))

    if request.method == "POST":
        f = request.form
        try:
            status_baru = f.get("status", "Aktif")
            # Audit Phase 4 — nilai lama status mahasiswa (Aktif/Cuti/
            # Non-Aktif/Drop Out), utk audit event kalau berubah. Ini
            # kolom lifecycle mahasiswa (Audit §23 Data Lifecycle) yang
            # relevan utk SPMI/AMI, beda dari status_ta yang sudah
            # ditangani workflow_ta tersendiri di Phase 3.
            status_lama = row["status"]
            # Audit P0 #2 — status_ta SENGAJA dikeluarkan dari UPDATE ini.
            # Sebelumnya operator bisa mengetik ulang status_ta lewat form
            # ini lepas dari histori pengajuan/pembimbing/seminar/sidang
            # yang sebenarnya (mis. set "LULUS" tanpa pernah ada baris
            # sidang LULUS). status_ta sekarang read-only di halaman ini —
            # nilainya murni hasil logic.recalculate_status_ta().
            conn.execute(
                "UPDATE mahasiswa SET nim=?, nama=?, jk=?, tempat_lahir=?, tgl_lahir=?, no_hp=?, "
                "email_nik=?, angkatan=?, konsentrasi=?, status=?, skema=?, catatan=? "
                "WHERE id=?",
                (
                    f.get("nim", "").strip(),
                    f.get("nama", "").strip(),
                    f.get("jk", ""),
                    f.get("tempat_lahir", ""),
                    f.get("tgl_lahir", ""),
                    f.get("no_hp", ""),
                    f.get("email_nik", ""),
                    f.get("angkatan", ""),
                    f.get("konsentrasi", ""),
                    status_baru,
                    f.get("skema", "Reguler"),
                    f.get("catatan", ""),
                    mid,
                ),
            )
            conn.commit()
            if status_lama != status_baru:
                _db.log(
                    conn, "Ubah Mahasiswa", f"{f.get('nim')} - {f.get('nama')}",
                    modul="Mahasiswa", entitas="Mahasiswa", entitas_id=mid,
                    nilai_lama=status_lama, nilai_baru=status_baru,
                )
            else:
                _db.log(
                    conn, "Ubah Mahasiswa", f"{f.get('nim')} - {f.get('nama')}",
                    modul="Mahasiswa", entitas="Mahasiswa", entitas_id=mid,
                )
            flash("Perubahan disimpan.", "ok")
            return redirect(url_for("mahasiswa.list_view"))
        except Exception as e:
            EH.flash_gagal_simpan(e, "Gagal menyimpan")

    # Audit Phase 3 — TA Workflow Engine: riwayat transisi status_ta
    # ditampilkan di halaman ini (read-only), pelengkap badge status_ta
    # yang sudah dibuat read-only sejak Phase 1 (P0 #2).
    riwayat_ta = []
    if row:
        from app import workflow_ta

        riwayat_ta = workflow_ta.riwayat_mahasiswa(conn, mid)

    return render_template(
        "mahasiswa_form.html",
        row=row,
        STATUS_TA_LIST=C.STATUS_TA_LIST,
        STATUS_MHS_LIST=C.STATUS_MHS_LIST,
        JK_LIST=C.JK_LIST,
        riwayat_ta=riwayat_ta,
    )


@bp.route("/<int:mid>/hapus", methods=["POST"])
def hapus(mid):
    conn = current_app.get_db()
    row = conn.execute("SELECT nim, nama FROM mahasiswa WHERE id=?", (mid,)).fetchone()
    if not row:
        flash("Data mahasiswa tidak ditemukan.", "error")
        return redirect(url_for("mahasiswa.list_view"))

    # Audit P0 #8 — SEBELUM perbaikan ini, tombol Hapus langsung
    # `DELETE FROM mahasiswa` tanpa pengecekan apa pun. Karena seminar,
    # sidang, yudisium, wisuda, tracer_study, dan penetapan_pembimbing
    # semuanya ON DELETE CASCADE ke mahasiswa(id), satu klik di sini
    # menghapus SELURUH histori akademik mahasiswa itu secara permanen —
    # termasuk riwayat yang sudah dipakai untuk rekap honor/SK yang mungkin
    # sudah terbit. Guard di bawah meniru pola yang sudah ada di
    # routes/pengaturan.py (hapus tahap_pengajuan): kalau masih ada histori
    # yang menempel, hapus DITOLAK dan operator diarahkan memakai field
    # "Status Mahasiswa" (Non-Aktif/Drop Out) sebagai gantinya — data lama
    # tetap utuh untuk akreditasi/audit, hanya disembunyikan dari daftar
    # aktif. Hard delete hanya diizinkan utk baris yang BENAR-BENAR belum
    # pernah punya histori apa pun (mis. salah input NIM lalu langsung
    # dihapus di hari yang sama).
    pemakaian = {
        "Pengajuan Judul": "pengajuan_judul",
        "SK Penetapan Pembimbing": "penetapan_pembimbing",
        "Seminar Proposal": "seminar",
        "Sidang Skripsi": "sidang",
        "Rencana Yudisium": "yudisium",
        "Wisuda": "wisuda",
        "Tracer Study": "tracer_study",
    }
    rincian = []
    for label, tabel in pemakaian.items():
        n = conn.execute(
            f"SELECT COUNT(*) n FROM {tabel} WHERE mahasiswa_id=?", (mid,)
        ).fetchone()["n"]
        if n:
            rincian.append(f"{label} ({n})")
    if rincian:
        flash(
            f'Mahasiswa "{row["nama"]}" tidak bisa dihapus permanen — masih punya histori di: '
            + ", ".join(rincian) +
            ". Menghapusnya akan menghilangkan riwayat akademik itu selamanya. "
            'Gunakan "Status Mahasiswa" = Non-Aktif atau Drop Out di halaman Ubah Data untuk '
            "menonaktifkan mahasiswa ini tanpa kehilangan riwayatnya.",
            "error",
        )
        return redirect(url_for("mahasiswa.list_view"))

    conn.execute("DELETE FROM mahasiswa WHERE id=?", (mid,))
    conn.commit()
    _db.log(
        conn, "Hapus Mahasiswa", f"{row['nim']} - {row['nama']}",
        modul="Mahasiswa", entitas="Mahasiswa", entitas_id=mid,
    )
    flash("Data mahasiswa dihapus.", "ok")
    return redirect(url_for("mahasiswa.list_view"))


@bp.route("/ekspor")
def ekspor():
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    conn = current_app.get_db()
    rows = conn.execute("SELECT * FROM mahasiswa ORDER BY nama").fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Mahasiswa"
    headers = [
        "NIM",
        "Nama",
        "JK",
        "Angkatan",
        "Skema",
        "Konsentrasi",
        "Status",
        "Status TA",
        "No HP",
        "Email/NIK",
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1E3A5F")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for r in rows:
        ws.append(
            [
                r["nim"],
                r["nama"],
                r["jk"],
                r["angkatan"],
                r["skema"] or "Reguler",
                r["konsentrasi"],
                r["status"],
                r["status_ta"],
                r["no_hp"],
                r["email_nik"],
            ]
        )
    for col in ws.columns:
        length = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(length + 2, 10), 45)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    _db.log(conn, "Ekspor Excel", "Data Mahasiswa")
    return send_file(
        buf,
        as_attachment=True,
        download_name="Data_Mahasiswa.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
