# -*- coding: utf-8 -*-
"""export_utils.py — Ekspor tabel ke file .xlsx (untuk dicetak / diarsipkan).

Audit Lanjutan (modul Kelulusan/Tracer Study): versi sebelumnya file ini
adalah sisa kode versi DESKTOP lama (import `tkinter`/`widgets`, memakai
`filedialog.asksaveasfilename` — dialog GUI yang tidak relevan sama sekali
untuk aplikasi web ini) dan TIDAK PERNAH dipakai oleh satu pun route Flask
(diverifikasi lewat pencarian `import export_utils` di seluruh app/routes/
— nihil). Kalau file itu sampai ter-import (mis. refactor ceroboh di masa
depan), aplikasi akan langsung crash karena modul `widgets` tidak ada di
paket ini dan `tkinter` belum tentu tersedia di server headless.

Sementara itu, fungsi `_kirim_excel()` yang SESUNGGUHNYA dipakai (di
routes/kelulusan.py dan routes/rekap.py) digandakan persis sama di 2 file
berbeda — pola klasik penyebab bug: kalau suatu saat perlu diperbaiki
(mis. batas lebar kolom, gaya header), gampang lupa mengubah salah satu
salinannya. File ini sekarang diisi ulang dengan implementasi TUNGGAL yang
sungguh dipakai kedua route tsb lewat `from app import export_utils`.
"""

import io

from flask import send_file


def kirim_excel(sheet_title, headers, rows):
    """Bangun 1 file .xlsx (1 sheet, header tebal + auto-width kolom) dari
    `headers` (list nama kolom) & `rows` (list tuple/list nilai per baris),
    lalu kirim sebagai attachment lewat Flask `send_file`.

    Dipakai oleh semua halaman "Ekspor Excel" yang sifatnya laporan 1
    sheet sederhana (Rencana Yudisium, Wisuda, Tracer Study Alumni, Rekap
    Pembimbing, RKP Seminar, dll). Laporan multi-sheet (mis. Rekap Program
    Kerja 2-sheet di routes/rekap.py) tetap membangun `openpyxl.Workbook()`
    sendiri karena butuh >1 sheet.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1E3A5F")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for r in rows:
        ws.append(list(r))
    for col in ws.columns:
        length = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(length + 2, 10), 45)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"{sheet_title.replace(' ', '_')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
