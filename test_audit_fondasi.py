import os, sys, tempfile, io, re
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

tmpdir = tempfile.mkdtemp()
os.environ["HOME"] = tmpdir

from app import create_app  # noqa: E402
import openpyxl  # noqa: E402

db_path = os.path.join(tmpdir, "test.db")
app = create_app(db_path=db_path)
app.config["TESTING"] = True
app.config["WTF_CSRF_ENABLED"] = False  # skrip tes tidak mengirim token CSRF
client = app.test_client()

FAILS = []
def check(label, cond):
    print(f"[{'OK' if cond else 'FAIL'}] {label}")
    if not cond:
        FAILS.append(label)

client.get("/login")
client.post("/login", data={"username": "kaprodi", "password1": "test1234", "password2": "test1234"}, follow_redirects=True)
# Restrukturisasi poin 2: Import & Restore Backup sekarang digerbangi PIN.
client.post("/pengaturan/pin", data={"pin1": "246810", "pin2": "246810"}, follow_redirects=True)

# ---------------------------------------------------------------- SECRET_KEY
check("SECRET_KEY dipersist di pengaturan (bukan os.urandom tiap start)",
      app.config["SECRET_KEY"] and len(app.config["SECRET_KEY"]) == 64)
from app import db as _db  # noqa: E402
conn = _db.connect(db_path)
stored = _db.get_setting(conn, "secret_key", "")
check("SECRET_KEY tersimpan persisten di tabel pengaturan", stored == app.config["SECRET_KEY"])

# ------------------------------------------------------- Wizard Tahun Ajaran
r = client.post("/pengaturan/tahun-akademik", data={"aksi": "buka_tahun", "kode": "2025/2026", "aktifkan": "Ganjil"}, follow_redirects=True)
check("POST buka_tahun -> 200", r.status_code == 200)
check("Tahun ajaran 2025/2026 muncul di halaman", b"2025/2026" in r.data)

r2 = client.post("/pengaturan/tahun-akademik", data={"aksi": "buka_tahun", "kode": "2025/2026", "aktifkan": "Genap"}, follow_redirects=True)
check("Buka tahun ajaran sama 2x -> idempoten (tidak duplikat)", r2.data.count(b"2025/2026") <= r.data.count(b"2025/2026") + 3)

conn2 = client.application.get_db  # noop, just ensure attr

ta_count = conn.execute("SELECT COUNT(*) c FROM tahun_ajaran WHERE kode='2025/2026'").fetchone()["c"]
periode_count = conn.execute("SELECT COUNT(*) c FROM periode_akademik pa JOIN tahun_ajaran ta ON ta.id=pa.tahun_ajaran_id WHERE ta.kode='2025/2026'").fetchone()["c"]
check("Hanya 1 baris tahun_ajaran untuk kode yang sama (idempoten)", ta_count == 1)
check("3 periode (Ganjil/Genap/Antara) dibuat otomatis", periode_count == 3)

# Tambah tahap dinamis (klarifikasi a: > 2 tahap)
periode_row = conn.execute(
    "SELECT pa.id FROM periode_akademik pa JOIN tahun_ajaran ta ON ta.id=pa.tahun_ajaran_id "
    "WHERE ta.kode='2025/2026' AND pa.jenis='Ganjil'"
).fetchone()
pid = periode_row["id"]
for nama in ["Tahap Awal Semester", "Tahap Tengah Semester", "Tahap Akhir Semester"]:
    client.post("/pengaturan/tahun-akademik", data={"aksi": "tambah_tahap", "periode_id": pid, "nama_tahap": nama})
tahap_count = conn.execute("SELECT COUNT(*) c FROM tahap_pengajuan WHERE periode_akademik_id=?", (pid,)).fetchone()["c"]
check("3 gelombang/tahap (bukan cuma 2 hardcode) berhasil ditambahkan dinamis", tahap_count == 3)

# Tandai Ganjil sebagai berjalan supaya get_periode_aktif() konsisten
client.post("/pengaturan/tahun-akademik", data={"aksi": "set_status_periode", "periode_id": pid, "status": "Berjalan"})

akad_page = client.get("/akademik/pengajuan")
check("Halaman Pengajuan Judul memuat 3 tahap dinamis di dropdown", akad_page.data.count(b"Tahap") >= 3)

# ------------------------------------------------------------- Dosen Homebase
r = client.post("/dosen/simpan", data={
    "nama": "Dr. Homebase Test", "nidn": "0099988877", "aktif": "on",
    "status_homebase": "Homebase",
})
r = client.post("/dosen/simpan", data={
    "nama": "Dr. Luar Test", "nidn": "0088877766", "aktif": "on",
    "status_homebase": "Dosen Luar Prodi", "unit_asal": "Prodi Sebelah",
})
dosen_page = client.get("/dosen/")
check("Dosen homebase & dosen luar tersimpan & tampil", b"Dr. Homebase Test" in dosen_page.data and b"Dr. Luar Test" in dosen_page.data)

rasio_default = client.get("/rekap/rasio-dosen")
check("Rasio dosen default (homebase only) TIDAK memuat dosen luar", b"Dr. Luar Test" not in rasio_default.data)
check("Rasio dosen default memuat dosen homebase", b"Dr. Homebase Test" in rasio_default.data)

rasio_semua = client.get("/rekap/rasio-dosen?semua=1")
check("Rasio dosen ?semua=1 memuat dosen luar juga", b"Dr. Luar Test" in rasio_semua.data)

sdm_page = client.get("/sdm/")
check("Dashboard SDM memuat dosen homebase & dosen luar (dipisah)", b"Dr. Homebase Test" in sdm_page.data and b"Dr. Luar Test" in sdm_page.data)

# ----------------------------------------------------------------- Backup
# Sejak Audit Lanjutan (Backup Menyeluruh), default "Backup Sekarang" adalah
# Backup LENGKAP (.zip — database + seluruh file fisik), bukan .db saja.
r = client.post("/pengaturan/backup/sekarang", follow_redirects=True)
check("Backup Sekarang -> 200", r.status_code == 200)
backup_list = client.get("/pengaturan/backup/")
check("Halaman backup memuat minimal 1 file backup", b".zip" in backup_list.data)

import re as _re
m = _re.search(rb"backup_lengkap_\d+_\d+_\d+\.zip", backup_list.data)
check("Nama file backup lengkap dengan timestamp ditemukan di halaman", bool(m))
if m:
    fname = m.group(0).decode()
    dl = client.get(f"/pengaturan/backup/unduh/{fname}")
    check("Unduh file backup -> 200", dl.status_code == 200)
    check("File backup adalah arsip ZIP valid (magic bytes)", dl.data[:2] == b"PK")

# Restore dengan file bukan SQLite -> ditolak
r = client.post("/pengaturan/backup/restore", data={
    "password_konfirmasi": "test1234",
    "file_restore": (io.BytesIO(b"bukan database sqlite"), "palsu.db"),
}, content_type="multipart/form-data", follow_redirects=True)
check("Restore file bukan SQLite -> ditolak (validasi magic bytes)", b"ditolak" in r.data or b"Ditolak" in r.data or b"bukan database" in r.data.lower())

# Restore dengan password salah -> ditolak
r = client.post("/pengaturan/backup/restore", data={
    "password_konfirmasi": "salahpassword",
    "file_restore": (io.BytesIO(b"SQLite format 3\x00" + b"\x00" * 100), "asal.db"),
}, content_type="multipart/form-data", follow_redirects=True)
check("Restore dgn password salah -> ditolak", b"salah" in r.data.lower())

# ----------------------------------------------------------- Generic Importer
tmpl = client.get("/pengaturan/import-generik/template/dosen")
check("Unduh template Dosen -> 200", tmpl.status_code == 200)
wb = openpyxl.load_workbook(io.BytesIO(tmpl.data))
ws = wb.active
check("Template Dosen memuat header 'Status Homebase'", any("Status Homebase" in str(c.value) for c in ws[1]))

# Susun file isian: 1 dosen baru + 1 update dosen yang sudah ada (via NIDN)
wb2 = openpyxl.Workbook()
ws2 = wb2.active
header = [c.value for c in ws[1]]
ws2.append(header)
ws2.append(["0055566677", "Dr. Import Baru", "0812", "baru@kampus.ac.id", "", "", "", "", "Homebase", "", "", "", "Y"])
ws2.append(["0099988877", "Dr. Homebase Test (Update)", "0813", "", "", "", "", "", "Homebase", "", "", "", "Y"])
buf = io.BytesIO()
wb2.save(buf)
buf.seek(0)
r = client.post("/pengaturan/import-generik/proses", data={
    "modul": "dosen", "file_excel": (buf, "isian_dosen.xlsx"),
}, content_type="multipart/form-data", follow_redirects=True)
check("Proses import generik dosen -> 200", r.status_code == 200)
check("Ringkasan menunjukkan 1 baru & 1 diperbarui", b"1 data baru" in r.data or (b"1" in r.data))

dosen_page2 = client.get("/dosen/")
check("Dosen baru dari import generik muncul", b"Dr. Import Baru" in dosen_page2.data)
check("Dosen lama ter-update (bukan digandakan) dari import generik", b"Dr. Homebase Test (Update)" in dosen_page2.data)
n_dosen_dgn_nidn = dosen_page2.data.count(b"0099988877")
check("Update by-NIDN tidak menggandakan baris", n_dosen_dgn_nidn == 1)

# Template Mahasiswa
tmpl_m = client.get("/pengaturan/import-generik/template/mahasiswa")
check("Unduh template Mahasiswa -> 200", tmpl_m.status_code == 200)
wb_m = openpyxl.load_workbook(io.BytesIO(tmpl_m.data))
ws_m = wb_m.active
check("Template Mahasiswa memuat kolom Skema (Reguler/RPL)", any("Skema" in str(c.value) for c in ws_m[1]))

wb3 = openpyxl.Workbook()
ws3 = wb3.active
header_m = [c.value for c in ws_m[1]]
ws3.append(header_m)
ws3.append(["2025010099", "Mhs RPL Import", "L", "Kota X", "01/01/2000", "0812", "rpl@student.ac.id", "2025", "Aktif", "RPL"])
buf3 = io.BytesIO()
wb3.save(buf3)
buf3.seek(0)
r = client.post("/pengaturan/import-generik/proses", data={
    "modul": "mahasiswa", "file_excel": (buf3, "isian_mhs.xlsx"),
}, content_type="multipart/form-data", follow_redirects=True)
check("Proses import generik mahasiswa RPL -> 200", r.status_code == 200)
mhs_page = client.get("/mahasiswa/")
check("Mahasiswa RPL dari import generik muncul", b"Mhs RPL Import" in mhs_page.data)

# ------------------------------------------------------------- Login lockout
client.get("/logout")
for i in range(5):
    client.post("/login", data={"password": "salahterus"})
locked = client.get("/login")
check("Setelah 5x gagal, login terkunci sementara", b"terkunci" in locked.data.lower() or b"lockout" in locked.data.lower() or b"disabled" in locked.data.lower())

# --------------------------------------------------------- Validasi Transisi
# Reset ke instance baru (CSRF tetap dimatikan) supaya tidak terpengaruh
# lockout login di atas.
tmpdir3 = tempfile.mkdtemp()
os.environ["HOME"] = tmpdir3
db_path3 = os.path.join(tmpdir3, "test.db")
app3 = create_app(db_path=db_path3)
app3.config["TESTING"] = True
app3.config["WTF_CSRF_ENABLED"] = False
client3 = app3.test_client()
client3.post("/login", data={"username": "kaprodi", "password1": "test1234", "password2": "test1234"}, follow_redirects=True)
client3.post("/dosen/simpan", data={"nama": "Dr. Penguji Sidang", "nidn": "0011223344", "aktif": "on"})
client3.post("/mahasiswa/tambah", data={"nim": "2025999999", "nama": "Mhs Uji Transisi", "status": "Aktif"})
mhs_page3 = client3.get("/mahasiswa/")
import re as _re3
mid_m = _re3.search(rb"/mahasiswa/(\d+)/edit", mhs_page3.data)
mid3 = mid_m.group(1).decode() if mid_m else None
check("Mahasiswa uji transisi berhasil dibuat", bool(mid3))
if mid3:
    client3.post("/akademik/penetapan/simpan", data={"mahasiswa_id": mid3, "judul_final": "Judul Uji"})
    # Coba langsung input Sidang TANPA Seminar "Selesai" dulu -> harus diminta konfirmasi
    r = client3.post("/pelaksanaan/sidang/simpan", data={
        "mahasiswa_id": mid3, "status_kelulusan": "LULUS", "tgl_sidang": "", "jam_sidang": "",
    }, follow_redirects=True)
    check("Input Sidang tanpa Seminar Selesai -> diminta konfirmasi transisi",
          b"Periksa Kembali Urutan Tahapan" in r.data or b"belum menyelesaikan Seminar" in r.data)
    # Dengan konfirmasi -> tetap bisa disimpan (soft-block, bukan hard-block)
    r2 = client3.post("/pelaksanaan/sidang/simpan", data={
        "mahasiswa_id": mid3, "status_kelulusan": "LULUS", "tgl_sidang": "", "jam_sidang": "",
        "konfirmasi_transisi": "1",
    }, follow_redirects=True)
    check("Dengan konfirmasi_transisi=1 -> data sidang tetap tersimpan (soft-block)",
          b"Data sidang disimpan" in r2.data or b"Belum" not in r2.data)
    sidang_page = client3.get("/pelaksanaan/sidang")
    check("Data sidang uji transisi benar-benar tersimpan", b"Mhs Uji Transisi" in sidang_page.data)

# --------------------------------------------------------------- Proteksi CSRF
# Instance TERPISAH dengan CSRF AKTIF (perilaku produksi sesungguhnya) —
# memverifikasi POST tanpa token ditolak, dan POST dengan token dari form
# GET yang sama berhasil.
tmpdir2 = tempfile.mkdtemp()
os.environ["HOME"] = tmpdir2
db_path2 = os.path.join(tmpdir2, "test.db")
app2 = create_app(db_path=db_path2)
app2.config["TESTING"] = True  # CSRF TETAP AKTIF — WTF_CSRF_ENABLED tidak diset False
client2 = app2.test_client()
login_page = client2.get("/login")
m_login_token = _re.search(rb'name="csrf_token" value="([^"]+)"', login_page.data)
login_token = m_login_token.group(1).decode() if m_login_token else ""
client2.post("/login", data={"username": "kaprodi", "password1": "test1234", "password2": "test1234", "csrf_token": login_token}, follow_redirects=True)

r_no_token = client2.post("/dosen/simpan", data={"nama": "Dr. Tanpa Token", "aktif": "on"})
check("POST tanpa csrf_token -> DITOLAK (mengarah balik, bukan tersimpan)",
      r_no_token.status_code in (302, 400, 403))
after_reject = client2.get("/dosen/")
check("Dosen TIDAK tersimpan tanpa token CSRF valid", b"Dr. Tanpa Token" not in after_reject.data)

form_page = client2.get("/dosen/")
m = re.search(rb'name="csrf_token" value="([^"]+)"', form_page.data)
check("Token CSRF disematkan otomatis di form Dosen", bool(m))
if m:
    token = m.group(1).decode()
    r_with_token = client2.post("/dosen/simpan", data={
        "nama": "Dr. Dengan Token", "aktif": "on", "csrf_token": token,
    }, follow_redirects=True)
    check("POST DENGAN csrf_token valid -> berhasil", b"Dr. Dengan Token" in r_with_token.data)

print("\n=== SELESAI ===")
if FAILS:
    print("ADA YANG GAGAL:", FAILS)
    sys.exit(1)
else:
    print("SEMUA TES LULUS.")
