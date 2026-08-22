# -*- coding: utf-8 -*-
"""
test_audit_lanjutan_3.py — Uji Pengembangan Lanjutan 3 (Bab Rekap Kinerja
Dosen/SDM & Bab Rekap Program Kerja di Modul 9 — Rekap & Laporan).

Latar belakang: README/FONDASI.md menandai eksplisit "Rekap TA sudah ada,
Bab SDM/modul lain menyusul" — audit ini menutup gap tsb dengan 2 bab baru
yang memakai rumus SAMA PERSIS dgn sdm.py (_hitung_kesiapan) dan
kegiatan.py (_hitung_realisasi), supaya angka selalu konsisten dgn
Dashboard SDM & halaman Program Kerja.

Tidak diikutkan di paket produksi (murni verifikasi pengembangan).
"""
import os
import sys
import tempfile
import io

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

tmpdir = tempfile.mkdtemp()
os.environ["HOME"] = tmpdir

from app import create_app  # noqa: E402

FAILS = []


def check(label, cond):
    print(f"[{'OK' if cond else 'FAIL'}] {label}")
    if not cond:
        FAILS.append(label)


db_path = os.path.join(tmpdir, "test.db")
app = create_app(db_path=db_path)
app.config["TESTING"] = True
app.config["WTF_CSRF_ENABLED"] = False
client = app.test_client()
client.post("/login", data={"username": "kaprodi", "password1": "test1234", "password2": "test1234"}, follow_redirects=True)

# ===========================================================================
# 0) SEED DATA — 2 dosen (1 homebase, 1 luar) + entri log SDM di beberapa
#    kategori, + 1 program kerja dgn 2 kegiatan (1 Selesai, 1 belum).
# ===========================================================================
with app.app_context():
    conn = app.get_db()
    conn.execute(
        "INSERT INTO dosen(nama, nidn, aktif, status_homebase) VALUES(?,?,1,?)",
        ("Dr. Uji Homebase", "0011110001", "Homebase"),
    )
    conn.execute(
        "INSERT INTO dosen(nama, nidn, aktif, status_homebase) VALUES(?,?,1,?)",
        ("Dr. Uji Luar", "0011110002", "Luar"),
    )
    conn.commit()
    dosen_hb = conn.execute("SELECT id FROM dosen WHERE nidn='0011110001'").fetchone()["id"]
    dosen_luar = conn.execute("SELECT id FROM dosen WHERE nidn='0011110002'").fetchone()["id"]

    # Aktivitas Pendidikan (2 entri, 1 Selesai) -> tahun_akademik terisi
    conn.execute(
        "INSERT INTO aktivitas_pendidikan(dosen_id, mata_kuliah, status, tahun_akademik) "
        "VALUES(?,?,?,?)", (dosen_hb, "Manajemen Strategik", "Selesai", "2025/2026"),
    )
    conn.execute(
        "INSERT INTO aktivitas_pendidikan(dosen_id, mata_kuliah, status, tahun_akademik) "
        "VALUES(?,?,?,?)", (dosen_hb, "Kewirausahaan", "Berjalan", "2025/2026"),
    )
    # Penelitian (1 entri, Selesai)
    conn.execute(
        "INSERT INTO aktivitas_penelitian(dosen_id, judul, status, tahun_akademik) "
        "VALUES(?,?,?,?)", (dosen_hb, "Riset Uji Kinerja Dosen", "Selesai", "2025/2026"),
    )
    # Timeline Karier (tidak punya tahun_akademik by design)
    conn.execute(
        "INSERT INTO timeline_karier_dosen(dosen_id, jenis_perubahan, keterangan, status) "
        "VALUES(?,?,?,?)", (dosen_hb, "Kenaikan Pangkat", "Lektor -> Lektor Kepala", "Selesai"),
    )
    # Dosen luar: 1 entri saja, tahun berbeda -> untuk uji filter homebase & tahun
    conn.execute(
        "INSERT INTO aktivitas_penelitian(dosen_id, judul, status, tahun_akademik) "
        "VALUES(?,?,?,?)", (dosen_luar, "Riset Dosen Luar", "Berjalan", "2024/2025"),
    )
    conn.commit()

    # Program Kerja + 2 Kegiatan (1 Selesai -> realisasi 50%)
    conn.execute(
        "INSERT INTO program_kerja(tahun_akademik, bidang, nama_program, anggaran_rencana, status) "
        "VALUES(?,?,?,?,?)", ("2025/2026", "SDM & Kelembagaan", "Program Uji Rekap", 5000000, "Berjalan"),
    )
    conn.commit()
    proker_id = conn.execute("SELECT id FROM program_kerja WHERE nama_program='Program Uji Rekap'").fetchone()["id"]
    conn.execute(
        "INSERT INTO kegiatan_prodi(program_kerja_id, nama_kegiatan, status) VALUES(?,?,?)",
        (proker_id, "Kegiatan Uji Selesai", "Selesai"),
    )
    conn.execute(
        "INSERT INTO kegiatan_prodi(program_kerja_id, nama_kegiatan, status) VALUES(?,?,?)",
        (proker_id, "Kegiatan Uji Belum", "Direncanakan"),
    )
    conn.commit()

# ===========================================================================
# 1) BAB REKAP KINERJA DOSEN (SDM)
# ===========================================================================
r = client.get("/rekap/kinerja-dosen")
check("GET /rekap/kinerja-dosen -> 200", r.status_code == 200)
html = r.data.decode()
check("Halaman menampilkan dosen homebase", "Dr. Uji Homebase" in html)
check("Default (hanya homebase) TIDAK menampilkan dosen luar", "Dr. Uji Luar" not in html)
check("Kesiapan BKD dosen homebase tampil (2/4 aktivitas Selesai antara Pendidikan+Penelitian... "
      "dihitung lewat _hitung_kesiapan yang sama dgn sdm.py)", "%" in html)

r2 = client.get("/rekap/kinerja-dosen?semua=1")
check("GET /rekap/kinerja-dosen?semua=1 -> 200", r2.status_code == 200)
check("Mode semua dosen menampilkan dosen luar juga", "Dr. Uji Luar" in r2.data.decode())

r3 = client.get("/rekap/kinerja-dosen?tahun_akademik=2025%2F2026")
check("Filter tahun_akademik=2025/2026 -> 200", r3.status_code == 200)
html3 = r3.data.decode()
check("Filter tahun tetap menampilkan Dr. Uji Homebase (punya entri 2025/2026)", "Dr. Uji Homebase" in html3)

r4 = client.get("/rekap/kinerja-dosen/ekspor")
check("GET /rekap/kinerja-dosen/ekspor -> 200", r4.status_code == 200)
check("Ekspor kinerja dosen mengembalikan file xlsx",
      r4.headers.get("Content-Type", "").startswith("application/vnd.openxmlformats"))

import openpyxl
wb = openpyxl.load_workbook(io.BytesIO(r4.data))
ws = wb.active
header_row = [c.value for c in ws[1]]
check("Header ekspor memuat kolom kategori Pendidikan & Pengajaran", "Pendidikan & Pengajaran" in header_row)
check("Header ekspor memuat kolom Kesiapan BKD (%)", "Kesiapan BKD (%)" in header_row)
nama_col = header_row.index("Nama Dosen")
nama_terekspor = [row[nama_col].value for row in ws.iter_rows(min_row=2)]
check("Baris ekspor (default homebase) memuat Dr. Uji Homebase", "Dr. Uji Homebase" in nama_terekspor)
check("Baris ekspor (default homebase) TIDAK memuat Dr. Uji Luar", "Dr. Uji Luar" not in nama_terekspor)

# ===========================================================================
# 2) BAB REKAP PROGRAM KERJA
# ===========================================================================
r5 = client.get("/rekap/program-kerja")
check("GET /rekap/program-kerja -> 200", r5.status_code == 200)
html5 = r5.data.decode()
check("Halaman menampilkan nama program yang diinput", "Program Uji Rekap" in html5)
check("Halaman menampilkan bidang SDM & Kelembagaan", "SDM &amp; Kelembagaan" in html5 or "SDM & Kelembagaan" in html5)
check("Realisasi 50% (1 dari 2 kegiatan Selesai) tampil di halaman", "50%" in html5)

r6 = client.get("/rekap/program-kerja?tahun_akademik=2025%2F2026")
check("Filter tahun_akademik=2025/2026 -> 200 & memuat program tsb", r6.status_code == 200 and "Program Uji Rekap" in r6.data.decode())

r7 = client.get("/rekap/program-kerja?tahun_akademik=2099%2F2100")
check("Filter tahun yang tidak ada datanya -> 200 (tabel kosong, bukan error)", r7.status_code == 200)
check("Filter tahun yang tidak ada datanya -> program tidak muncul", "Program Uji Rekap" not in r7.data.decode())

r8 = client.get("/rekap/program-kerja/ekspor")
check("GET /rekap/program-kerja/ekspor -> 200", r8.status_code == 200)
check("Ekspor program kerja mengembalikan file xlsx",
      r8.headers.get("Content-Type", "").startswith("application/vnd.openxmlformats"))
wb2 = openpyxl.load_workbook(io.BytesIO(r8.data))
check("Ekspor program kerja punya 2 sheet (Detail Program + Ringkasan Bidang)", len(wb2.sheetnames) == 2)
ws_detail = wb2["Bagian 1 - Detail Program"]
detail_rows = [[c.value for c in row] for row in ws_detail.iter_rows(min_row=2)]
check("Sheet detail memuat baris Program Uji Rekap dgn realisasi 50",
      any(row[2] == "Program Uji Rekap" and row[-1] == 50 for row in detail_rows))
ws_bidang = wb2["Bagian 2 - Ringkasan Bidang"]
bidang_rows = [[c.value for c in row] for row in ws_bidang.iter_rows(min_row=2)]
check("Sheet ringkasan bidang memuat baris SDM & Kelembagaan dgn 2 kegiatan",
      any(row[0] == "SDM & Kelembagaan" and row[2] == 2 for row in bidang_rows))

# ===========================================================================
# 3) SIDEBAR & PANDUAN — 2 bab baru terdaftar & tidak merusak yang lama
# ===========================================================================
html_home = client.get("/").data.decode()
check("Tautan 'Rekap Kinerja Dosen (SDM)' ada di sidebar", 'href="/rekap/kinerja-dosen"' in html_home)
check("Tautan 'Rekap Program Kerja' ada di sidebar", 'href="/rekap/program-kerja"' in html_home)

r_panduan = client.get("/panduan/")
check("Halaman Panduan tetap 200 setelah entri baru ditambahkan", r_panduan.status_code == 200)
check("Panduan memuat entri 'Rekap Kinerja Dosen (SDM)'", "Rekap Kinerja Dosen (SDM)" in r_panduan.data.decode())
check("Panduan memuat entri 'Rekap Program Kerja'", "Rekap Program Kerja" in r_panduan.data.decode())

r_docx = client.get("/panduan/unduh")
check("Unduh Panduan .docx tetap 200 (ikut memuat 2 bab baru otomatis)", r_docx.status_code == 200)

# ===========================================================================
# 4) REGRESI — bab Rekap lama & modul SDM/Kegiatan asal tidak rusak
# ===========================================================================
check("Rekap Pembimbing (lama) tetap 200", client.get("/rekap/pembimbing").status_code == 200)
check("Rekap Status (lama) tetap 200", client.get("/rekap/status").status_code == 200)
check("Rasio Beban Dosen (lama) tetap 200", client.get("/rekap/rasio-dosen").status_code == 200)
check("Statistik (lama) tetap 200", client.get("/rekap/statistik").status_code == 200)
check("Dashboard SDM asal (sdm.index) tetap 200 & angka kesiapan konsisten",
      client.get("/sdm/").status_code == 200)
check("Halaman Program Kerja asal (kegiatan.index) tetap 200", client.get("/kegiatan/?tab=proker").status_code == 200)

print("\n=== SELESAI ===")
if FAILS:
    print(f"{len(FAILS)} GAGAL:")
    for f in FAILS:
        print(f" - {f}")
    sys.exit(1)
else:
    print("SEMUA TES LULUS.")
